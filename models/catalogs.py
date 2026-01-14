"""Funciones para cargar catálogos y archivos masivos."""

from __future__ import annotations

import csv
import os
import re
from datetime import date, datetime
from typing import Dict, Iterable, Iterator, List, Sequence, Tuple

from settings import BASE_DIR, DETAIL_LOOKUP_ALIASES

CSV_IMPORT_ENCODINGS: tuple[str, ...] = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
EXCEL_IMPORT_EXTENSIONS: tuple[str, ...] = (".xlsx", ".xls")


def normalize_detail_catalog_key(key: str) -> str:
    """Normaliza una clave de catálogo de detalle a minúsculas sin espacios."""

    return (key or "").strip().lower()


def _normalize_identifier_token(value: str) -> str:
    """Genera un token comparable eliminando separadores y acentos básicos."""

    return re.sub(r"[^a-z0-9]", "", normalize_detail_catalog_key(value))


def _resolve_id_fields(entity_name: str, fieldnames: list[str] | None) -> tuple[str | None, str | None]:
    """Selecciona la columna llave incluso si está desordenada o usa alias.

    Retorna el nombre original de la columna encontrada y el identificador
    canónico (por ejemplo ``id_colaborador``) cuando es posible deducirlo.
    """

    alias_token_map = {
        _normalize_identifier_token(name): canonical
        for canonical, aliases in (DETAIL_LOOKUP_ALIASES or {}).items()
        for name in tuple(aliases or ()) + (canonical,)
    }
    normalized_entity = _normalize_identifier_token(entity_name)
    canonical_field = alias_token_map.get(normalized_entity)
    tokens = {token for token, target in alias_token_map.items() if target == canonical_field}
    if not tokens:
        tokens = set(alias_token_map.keys())

    chosen_field = None
    for field in fieldnames or []:
        normalized_field = _normalize_identifier_token(field)
        starts_with_id = normalize_detail_catalog_key(field).startswith("id")
        if normalized_field in tokens or starts_with_id:
            chosen_field = field
            canonical_field = canonical_field or alias_token_map.get(normalized_field)
            break

    if not canonical_field and chosen_field:
        canonical_field = normalize_detail_catalog_key(chosen_field)
    return chosen_field, canonical_field


def load_detail_catalogs(base_dir: str | os.PathLike = BASE_DIR) -> Dict[str, Dict[str, Dict[str, str]]]:
    """Lee todos los archivos ``*_details.csv`` disponibles en la carpeta base.

    Los valores se leen como texto para preservar ceros a la izquierda y evitar
    conversiones implícitas de tipo.
    """

    base_dir = os.fspath(base_dir)
    catalogs: Dict[str, Dict[str, Dict[str, str]]] = {}
    try:
        filenames = [
            name
            for name in os.listdir(base_dir)
            if name.lower().endswith("details.csv")
        ]
    except OSError:
        return catalogs

    for filename in filenames:
        path = os.path.join(base_dir, filename)
        entity_name = filename[:-len("_details.csv")].lower()
        try:
            rows, fieldnames, _encoding = _read_csv_rows_with_fallback(path)
        except (FileNotFoundError, OSError):
            continue
        key_field, canonical_id_field = _resolve_id_fields(entity_name, fieldnames)
        if not key_field:
            continue
        for row in rows:
            key = (row.get(key_field) or "").strip()
            if not key:
                continue
            cleaned_row = {(k or ""): (v or "").strip() for k, v in row.items()}
            canonical_id_field = canonical_id_field or key_field
            if canonical_id_field and canonical_id_field not in cleaned_row:
                cleaned_row[canonical_id_field] = key
            catalogs.setdefault(entity_name, {})[key] = cleaned_row
    return catalogs


def build_detail_catalog_id_index(
    catalogs: Dict[str, Dict[str, Dict[str, str]]]
) -> Dict[str, Dict[str, Dict[str, str]]]:
    """Construye un índice por columnas ``id_*`` reutilizando los catálogos cargados."""

    if not catalogs:
        return {}

    index: Dict[str, Dict[str, Dict[str, str]]] = {}
    for canonical_key, aliases in (DETAIL_LOOKUP_ALIASES or {}).items():
        normalized_canonical = normalize_detail_catalog_key(canonical_key)
        candidate_keys = [normalized_canonical]
        candidate_keys.extend(
            normalize_detail_catalog_key(alias)
            for alias in aliases or ()
        )
        seen = set()
        for candidate in candidate_keys:
            if not candidate or candidate in seen:
                continue
            seen.add(candidate)
            lookup = catalogs.get(candidate)
            if lookup is not None:
                index[normalized_canonical] = lookup
                break
    return index


def iter_massive_csv_rows(filename: str) -> Iterator[Dict[str, str]]:
    """Itera sobre los CSV masivos eliminando filas vacías y espacios extra."""

    rows, _fieldnames, _encoding = _read_csv_rows_with_fallback(filename)
    for row in rows:
        cleaned: Dict[str, str] = {}
        for key, value in row.items():
            if key is None:
                continue
            key = key.strip()
            if isinstance(value, str):
                value = value.strip()
            cleaned[key] = value
        if cleaned:
            yield cleaned


def iter_massive_rows(filename: str | os.PathLike) -> Iterator[Dict[str, str]]:
    """Itera sobre archivos masivos CSV o Excel normalizando valores como texto."""

    filename = os.fspath(filename)
    if _is_excel_file(filename):
        yield from _iter_excel_rows(filename)
        return
    yield from iter_massive_csv_rows(filename)


def read_import_headers(
    filename: str | os.PathLike,
    *,
    encodings: Sequence[str] | None = None,
) -> list[str]:
    """Lee encabezados desde CSV o Excel usando la misma lógica de importación."""

    filename = os.fspath(filename)
    if _is_excel_file(filename):
        return _read_excel_headers(filename)
    return read_csv_headers_with_fallback(filename, encodings=encodings)


def read_csv_headers_with_fallback(
    filename: str | os.PathLike,
    *,
    encodings: Sequence[str] | None = None,
) -> list[str]:
    """Lee encabezados CSV intentando múltiples codificaciones."""

    encodings = tuple(encodings or CSV_IMPORT_ENCODINGS)
    errors: list[str] = []
    for encoding in encodings:
        try:
            with open(filename, newline="", encoding=encoding) as handle:
                reader = csv.DictReader(line for line in handle if line.strip())
                return list(reader.fieldnames or [])
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
            continue
    error_message = (
        "No se pudo leer el archivo con las codificaciones "
        f"{', '.join(encodings)}. Detalles: {'; '.join(errors)}"
    )
    raise ValueError(error_message)


def read_csv_rows_with_fallback(
    filename: str | os.PathLike,
    *,
    encodings: Sequence[str] | None = None,
) -> tuple[list[dict[str, str]], list[str], str]:
    """Lee todas las filas CSV asegurando codificación compatible."""

    return _read_csv_rows_with_fallback(filename, encodings=encodings)


def parse_involvement_entries(raw_value: str | Iterable[str]) -> List[Tuple[str, str]]:
    """Parsea la columna de involucramientos proveniente del CSV combinado."""

    if not raw_value:
        return []
    entries: List[Tuple[str, str]] = []
    if isinstance(raw_value, (list, tuple)):
        raw_value = ";".join(raw_value)
    for chunk in str(raw_value).split(';'):
        chunk = chunk.strip()
        if not chunk:
            continue
        if ':' in chunk:
            collaborator, amount = chunk.split(':', 1)
        else:
            collaborator, amount = chunk, ''
        collaborator = collaborator.strip()
        amount = amount.strip()
        if collaborator:
            entries.append((collaborator, amount))
    return entries


__all__ = [
    "build_detail_catalog_id_index",
    "CSV_IMPORT_ENCODINGS",
    "EXCEL_IMPORT_EXTENSIONS",
    "iter_massive_csv_rows",
    "iter_massive_rows",
    "load_detail_catalogs",
    "normalize_detail_catalog_key",
    "parse_involvement_entries",
    "read_import_headers",
    "read_csv_headers_with_fallback",
    "read_csv_rows_with_fallback",
]


def _read_csv_rows_with_fallback(
    filename: str | os.PathLike,
    *,
    encodings: Sequence[str] | None = None,
) -> tuple[list[dict[str, str]], list[str], str]:
    encodings = tuple(encodings or CSV_IMPORT_ENCODINGS)
    errors: list[str] = []
    for encoding in encodings:
        try:
            with open(filename, newline="", encoding=encoding) as handle:
                reader = csv.DictReader(line for line in handle if line.strip())
                rows = list(reader)
                fieldnames = list(reader.fieldnames or [])
            return rows, fieldnames, encoding
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
            continue
    error_message = (
        "No se pudo leer el archivo con las codificaciones "
        f"{', '.join(encodings)}. Detalles: {'; '.join(errors)}"
    )
    raise ValueError(error_message)


def _is_excel_file(filename: str) -> bool:
    return os.path.splitext(filename.lower())[1] in EXCEL_IMPORT_EXTENSIONS


def _excel_cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _iter_excel_rows(filename: str) -> Iterator[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencias de entorno
        raise ValueError("openpyxl es requerido para importar archivos Excel.") from exc

    workbook = load_workbook(filename, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return
        headers = [
            (str(cell).strip() if cell is not None else "")
            for cell in header_row
        ]
        for row in rows:
            cleaned: Dict[str, str] = {}
            for idx, value in enumerate(row or ()):
                if idx >= len(headers):
                    continue
                header = headers[idx]
                if not header:
                    continue
                cell_text = _excel_cell_to_text(value)
                if isinstance(cell_text, str):
                    cell_text = cell_text.strip()
                cleaned[header] = cell_text
            if cleaned:
                yield cleaned
    finally:
        workbook.close()


def _read_excel_headers(filename: str) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencias de entorno
        raise ValueError("openpyxl es requerido para importar archivos Excel.") from exc

    workbook = load_workbook(filename, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []
        return [
            (str(cell).strip() if cell is not None else "")
            for cell in header_row
        ]
    finally:
        workbook.close()
