"""Genera una versión mejorada del wireframe de formularios en Excel.

Este script actualiza las hojas del archivo
``docs/formulario_investigaciones_wireframe.xlsx`` con la estructura real de la
UI, los catálogos y los exportes definidos en el código. Se apoya en los
diagramas y definiciones de ``docs/`` y ``settings.py`` para mantener las
descripciones alineadas con las reglas de negocio del Design document CM.pdf.

Uso:
    python tools/generate_improved_wireframe_excel.py \\
        --output docs/formulario_investigaciones_wireframe_mejorado.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _ensure_repo_root_on_path() -> Path:
    """Asegura que el root del repositorio esté disponible en sys.path.

    Esto evita errores de importación cuando el script se ejecuta desde fuera
    del directorio raíz (por ejemplo, lanzándolo desde ``tools/``).
    """

    repo_root = Path(__file__).resolve().parents[1]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))
    return repo_root


_ensure_repo_root_on_path()

from app import EXPORT_HEADERS  # noqa: E402
from settings import EVENTOS_HEADER_CANONICO, EVENTOS_HEADER_CANONICO_START  # noqa: E402
from validators import LOG_FIELDNAMES  # noqa: E402


HEADER_FONT = Font(bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)

CLASSIFY_MANUAL = "Manual"
CLASSIFY_LIST = "Lista"
CLASSIFY_LAKE = "Lake"
CLASSIFY_SYSTEM = "Sistema"
CLASSIFY_UI = "UI"
CLASSIFY_EXPORT = "Exportación"
CLASSIFY_REPORT = "Reporte"
CLASSIFY_VALIDATION = "Validación"
CLASSIFY_MAPPING = "Mapeo"


EXPORT_FIELD_DESCRIPTIONS = {
    "casos.csv": {
        "id_caso": "Identificador único del caso (AAAA-XXXX).",
        "tipo_informe": "Tipo de informe a generar (Gerencia/Interno/Credicorp).",
        "categoria1": "Categoría nivel 1 del caso (taxonomía).",
        "categoria2": "Categoría nivel 2 del caso (taxonomía).",
        "modalidad": "Modalidad del caso (taxonomía).",
        "canal": "Canal donde ocurrió el evento.",
        "proceso": "Proceso impactado.",
        "fecha_de_ocurrencia": "Fecha de ocurrencia del caso (YYYY-MM-DD).",
        "fecha_de_descubrimiento": "Fecha de descubrimiento del caso (YYYY-MM-DD).",
        "centro_costos": "Centros de costos separados por ';'.",
        "matricula_investigador": "Matrícula/ID del investigador principal.",
        "investigador_nombre": "Nombre y apellidos del investigador principal.",
        "investigador_cargo": "Cargo del investigador principal.",
    },
    "clientes.csv": {
        "id_cliente": "Identificador del cliente según tipo de documento.",
        "id_caso": "Caso asociado.",
        "nombres": "Nombres del cliente.",
        "apellidos": "Apellidos del cliente.",
        "tipo_id": "Tipo de documento (DNI/RUC/Pasaporte/etc.).",
        "flag": "Rol del cliente (Involucrado/Afectado/No aplica).",
        "telefonos": "Teléfonos separados por ';'.",
        "correos": "Correos separados por ';'.",
        "direcciones": "Direcciones físicas separadas por ';'.",
        "accionado": "Tribus/equipos accionados (lista múltiple).",
    },
    "colaboradores.csv": {
        "id_colaborador": "ID del Team Member (letra + 5 dígitos).",
        "id_caso": "Caso asociado.",
        "flag": "Rol del colaborador (Involucrado/Relacionado/etc.).",
        "nombres": "Nombres del colaborador.",
        "apellidos": "Apellidos del colaborador.",
        "division": "División del colaborador.",
        "area": "Área del colaborador.",
        "servicio": "Servicio del colaborador.",
        "puesto": "Puesto del colaborador.",
        "fecha_carta_inmediatez": "Fecha de carta de inmediatez (YYYY-MM-DD).",
        "fecha_carta_renuncia": "Fecha de renuncia/cese (YYYY-MM-DD).",
        "nombre_agencia": "Nombre de agencia (condicional).",
        "codigo_agencia": "Código de agencia (6 dígitos, condicional).",
        "tipo_falta": "Tipo de falta.",
        "tipo_sancion": "Tipo de sanción.",
    },
    "productos.csv": {
        "id_producto": "Identificador del producto investigado.",
        "id_caso": "Caso asociado.",
        "id_cliente": "Cliente titular del producto.",
        "categoria1": "Categoría nivel 1 (override por producto).",
        "categoria2": "Categoría nivel 2 (override por producto).",
        "modalidad": "Modalidad (override por producto).",
        "canal": "Canal (override por producto).",
        "proceso": "Proceso impactado (override por producto).",
        "fecha_ocurrencia": "Fecha de ocurrencia del producto (YYYY-MM-DD).",
        "fecha_descubrimiento": "Fecha de descubrimiento del producto (YYYY-MM-DD).",
        "monto_investigado": "Monto investigado total.",
        "tipo_moneda": "Moneda principal.",
        "monto_perdida_fraude": "Monto de pérdida por fraude.",
        "monto_falla_procesos": "Monto por falla de procesos.",
        "monto_contingencia": "Monto por contingencia.",
        "monto_recuperado": "Monto recuperado.",
        "monto_pago_deuda": "Monto de pago de deuda.",
        "tipo_producto": "Tipo de producto (tarjeta, crédito, etc.).",
    },
    "producto_reclamo.csv": {
        "id_reclamo": "ID del reclamo (C########).",
        "id_caso": "Caso asociado.",
        "id_producto": "Producto asociado.",
        "nombre_analitica": "Nombre de analítica contable.",
        "codigo_analitica": "Código de analítica contable.",
    },
    "involucramiento.csv": {
        "id_producto": "Producto asociado.",
        "id_caso": "Caso asociado.",
        "tipo_involucrado": "Tipo de involucrado (cliente/colaborador).",
        "id_colaborador": "ID del colaborador involucrado.",
        "id_cliente_involucrado": "ID del cliente involucrado.",
        "monto_asignado": "Monto asignado al involucrado.",
    },
    "detalles_riesgo.csv": {
        "id_riesgo": "Identificador del riesgo.",
        "id_caso": "Caso asociado.",
        "lider": "Responsable del riesgo.",
        "descripcion": "Descripción del riesgo.",
        "criticidad": "Criticidad del riesgo.",
        "exposicion_residual": "Exposición residual del riesgo.",
        "planes_accion": "Planes de acción separados por ';'.",
    },
    "detalles_norma.csv": {
        "id_norma": "Identificador de norma (XXXX.XXX.XX.XX).",
        "id_caso": "Caso asociado.",
        "descripcion": "Descripción de la norma.",
        "fecha_vigencia": "Fecha de vigencia (YYYY-MM-DD).",
        "acapite_inciso": "Acápite/inciso relacionado.",
        "detalle_norma": "Detalle de la norma transgredida.",
    },
    "analisis.csv": {
        "id_caso": "Caso asociado.",
        "antecedentes": "Antecedentes del caso.",
        "modus_operandi": "Modus operandi del evento.",
        "hallazgos": "Hallazgos principales.",
        "descargos": "Descargos del colaborador.",
        "conclusiones": "Conclusiones generales.",
        "recomendaciones": "Recomendaciones y mejoras.",
        "comentario_breve": "Resumen breve sin saltos de línea (máx. 150).",
        "comentario_amplio": "Resumen amplio sin saltos de línea (máx. 750).",
    },
}


def _apply_headers(sheet, headers: Sequence[str]) -> None:
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def _append_rows(sheet, rows: Iterable[Sequence[str | None]]) -> None:
    for row in rows:
        sheet.append(list(row))
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = WRAP


def _section_row(title: str, *, columns: int) -> list[str | None]:
    return [title] + [""] * (columns - 1)


def _normalize_text(value: str | None) -> str:
    return (value or "").strip().lower()


def _classify_input(field_type: str | None, source: str | None, description: str | None = None) -> str:
    normalized_type = _normalize_text(field_type)
    normalized_source = _normalize_text(source)
    normalized_desc = _normalize_text(description)

    if not normalized_type:
        return ""
    if all(not text for text in (normalized_source, normalized_desc)) and normalized_type in {"tabla", "sección"}:
        return CLASSIFY_SYSTEM
    if "botón" in normalized_type or "button" in normalized_type:
        return CLASSIFY_UI
    if "progress" in normalized_type:
        return CLASSIFY_SYSTEM
    if "tabla" in normalized_type or normalized_type == "tabla":
        return CLASSIFY_SYSTEM
    if "label" in normalized_type:
        if "solo lectura" in normalized_type and ".csv" in normalized_source:
            return CLASSIFY_LAKE
        return CLASSIFY_SYSTEM
    if "checkbutton" in normalized_type:
        return CLASSIFY_MANUAL
    if "listbox" in normalized_type or "combobox" in normalized_type:
        if ".csv" in normalized_source or "details.csv" in normalized_source:
            return CLASSIFY_LAKE
        if "settings." in normalized_source or "catálogo" in normalized_source:
            return CLASSIFY_LIST
        return CLASSIFY_LIST
    if "selector de fecha" in normalized_type:
        if ".csv" in normalized_source:
            return CLASSIFY_LAKE
        return CLASSIFY_MANUAL
    if "entry" in normalized_type or "texto enriquecido" in normalized_type:
        if "solo lectura" in normalized_type and ".csv" in normalized_source:
            return CLASSIFY_LAKE
        if ".csv" in normalized_source:
            return CLASSIFY_LAKE
        if "resumen" in normalized_desc or "auto" in normalized_desc:
            return CLASSIFY_SYSTEM
        return CLASSIFY_MANUAL
    return CLASSIFY_SYSTEM


def _append_classification(
    rows: list[Sequence[str | None]],
    *,
    field_type_index: int,
    source_index: int,
    description_index: int,
    section_label: str = CLASSIFY_SYSTEM,
) -> list[Sequence[str | None]]:
    updated: list[Sequence[str | None]] = []
    for row in rows:
        row_values = list(row)
        if len(row_values) > 1 and all(not str(cell or "").strip() for cell in row_values[1:]):
            row_values.append(section_label)
            updated.append(tuple(row_values))
            continue
        field_type = row_values[field_type_index] if len(row_values) > field_type_index else ""
        source = row_values[source_index] if len(row_values) > source_index else ""
        description = row_values[description_index] if len(row_values) > description_index else ""
        row_values.append(_classify_input(field_type, source, description))
        updated.append(tuple(row_values))
    return updated


def _case_and_participants_rows() -> list[Sequence[str | None]]:
    headers = ("Campo", "Tipo", "Descripción", "Autocompletado / Fuente", "Clasificación")
    rows: list[Sequence[str | None]] = []

    rows.append(_section_row("1. Datos generales del caso", columns=len(headers)))
    rows.extend(
        [
            (
                "Número de caso (AAAA-XXXX)",
                "Entry (texto)",
                "Identificador único del caso. Se valida formato AAAA-XXXX y se usa en la llave técnica.",
                "No",
            ),
            (
                "Tipo de informe",
                "Combobox",
                "Tipo de reporte a generar (Gerencia/Interno/Credicorp).",
                "settings.TIPO_INFORME_LIST",
            ),
            (
                "Categoría nivel 1",
                "Combobox",
                "Categoría principal del fraude. Controla la lista de categoría nivel 2.",
                "settings.TAXONOMIA",
            ),
            (
                "Categoría nivel 2",
                "Combobox",
                "Subcategoría dependiente de la categoría nivel 1.",
                "settings.TAXONOMIA",
            ),
            (
                "Modalidad",
                "Combobox",
                "Modalidad específica dependiente de categoría 1/2.",
                "settings.TAXONOMIA",
            ),
            (
                "Canal",
                "Combobox",
                "Canal donde ocurrió el evento. Puede autocompletarse al elegir el ID de proceso.",
                "settings.CANAL_LIST / process_details.csv",
            ),
            (
                "ID Proceso",
                "Entry (texto)",
                "Identificador del proceso impactado (BPID-XXXXXX o BPID-RNF-XXXXXX).",
                "process_details.csv (autopobla canal/proceso)",
            ),
            (
                "Botón \"Seleccionar\" (ID Proceso)",
                "Botón",
                "Abre el selector de procesos para escoger un ID válido.",
                "Catálogo de procesos",
            ),
            (
                "Proceso impactado",
                "Combobox",
                "Proceso del negocio afectado por el evento.",
                "settings.PROCESO_LIST / process_details.csv",
            ),
            (
                "Centro de costos del caso (; separados)",
                "Entry (texto)",
                "Lista de centros de costos. Cada valor debe ser numérico y >=5 dígitos.",
                "No",
            ),
            (
                "Ocurrencia (YYYY-MM-DD)",
                "Selector de fecha",
                "Fecha de ocurrencia del caso. Debe ser <= hoy y anterior al descubrimiento.",
                "No",
            ),
            (
                "Descubrimiento (YYYY-MM-DD)",
                "Selector de fecha",
                "Fecha de descubrimiento. Debe ser <= hoy y posterior a la ocurrencia.",
                "No",
            ),
            (
                "Matrícula investigador",
                "Entry (texto)",
                "ID del investigador principal (letra + 5 dígitos) para autocompletar datos.",
                "team_details.csv",
            ),
            (
                "Nombre y apellidos",
                "Entry (solo lectura)",
                "Nombre del investigador autocompletado desde catálogos.",
                "team_details.csv",
            ),
            (
                "Puesto",
                "Label (solo lectura)",
                "Cargo del investigador autocompletado.",
                "team_details.csv",
            ),
        ]
    )

    rows.append(_section_row("2. Clientes implicados", columns=len(headers)))
    rows.extend(
        [
            (
                "Tabla resumen de clientes",
                "Tabla",
                "Resumen de clientes registrados en el caso.",
                "Resumen dinámico",
            ),
            ("Botón \"Añadir cliente\"", "Botón", "Crea un nuevo cliente.", ""),
            (
                "Afectación interna",
                "Checkbutton",
                "Marca si el cliente pertenece a afectación interna.",
                "No",
            ),
            (
                "Tipo de ID",
                "Combobox",
                "Tipo de documento del cliente (DNI/RUC/Pasaporte/etc.).",
                "settings.TIPO_ID_LIST",
            ),
            (
                "ID del cliente",
                "Entry (texto)",
                "Documento del cliente. Valida longitud según tipo de ID.",
                "No",
            ),
            ("Nombres", "Entry (texto)", "Nombres del cliente.", "client_details.csv"),
            ("Apellidos", "Entry (texto)", "Apellidos del cliente.", "client_details.csv"),
            (
                "Flag",
                "Combobox",
                "Rol del cliente (Involucrado/Afectado/No aplica).",
                "settings.FLAG_CLIENTE_LIST",
            ),
            (
                "Accionado (lista múltiple)",
                "Listbox",
                "Tribus/equipos accionados. Valida selección múltiple.",
                "settings.ACCIONADO_OPTIONS",
            ),
            (
                "Teléfonos (separados por ;)",
                "Entry (texto)",
                "Teléfonos del cliente. Validación de números (+, 6-15 dígitos).",
                "client_details.csv",
            ),
            (
                "Correos (separados por ;)",
                "Entry (texto)",
                "Correos del cliente. Deben cumplir formato de email.",
                "client_details.csv",
            ),
            (
                "Direcciones (separados por ;)",
                "Entry (texto)",
                "Direcciones físicas del cliente (opcional).",
                "client_details.csv",
            ),
            ("Botón \"Eliminar cliente\"", "Botón", "Elimina el cliente del caso.", ""),
        ]
    )

    rows.append(_section_row("3. Productos investigados", columns=len(headers)))
    rows.extend(
        [
            ("Tabla resumen de productos", "Tabla", "Resumen de productos investigados.", ""),
            ("Botón \"Agregar producto\"", "Botón", "Crea un nuevo producto.", ""),
            (
                "ID del producto",
                "Entry (texto)",
                "Identificador del producto. Valida longitud según tipo de producto.",
                "product_details.csv (autopobla atributos)",
            ),
            (
                "Cliente",
                "Combobox",
                "Cliente titular del producto.",
                "Lista de clientes registrados / product_details.csv",
            ),
            (
                "Categoría 1",
                "Combobox",
                "Categoría principal del riesgo (override por producto).",
                "settings.TAXONOMIA / product_details.csv",
            ),
            (
                "Categoría 2",
                "Combobox",
                "Subcategoría del producto (override).",
                "settings.TAXONOMIA / product_details.csv",
            ),
            (
                "Modalidad",
                "Combobox",
                "Modalidad específica (override).",
                "settings.TAXONOMIA / product_details.csv",
            ),
            ("Canal", "Combobox", "Canal del evento (override).", "settings.CANAL_LIST / product_details.csv"),
            ("Proceso", "Combobox", "Proceso impactado (override).", "settings.PROCESO_LIST / product_details.csv"),
            (
                "Tipo de producto",
                "Combobox",
                "Clasificación comercial (tarjeta, crédito, etc.).",
                "settings.TIPO_PRODUCTO_LIST / product_details.csv",
            ),
            (
                "Fecha de ocurrencia (YYYY-MM-DD)",
                "Selector de fecha",
                "Fecha del evento. Requerida; debe ser <= hoy y < descubrimiento.",
                "product_details.csv",
            ),
            (
                "Fecha de descubrimiento (YYYY-MM-DD)",
                "Selector de fecha",
                "Fecha en que se detecta el evento. Requerida; <= hoy y > ocurrencia.",
                "product_details.csv",
            ),
            (
                "Monto investigado",
                "Entry (número)",
                "Monto total investigado. Debe ser suma de pérdida+falla+contingencia+recuperado.",
                "product_details.csv",
            ),
            ("Moneda", "Combobox", "Moneda principal.", "settings.TIPO_MONEDA_LIST / product_details.csv"),
            (
                "Monto pérdida fraude",
                "Entry (número)",
                "Monto de pérdida por fraude. >=0, 12 dígitos, 2 decimales.",
                "product_details.csv",
            ),
            (
                "Monto falla procesos",
                "Entry (número)",
                "Monto por falla de procesos. >=0, 12 dígitos, 2 decimales.",
                "product_details.csv",
            ),
            (
                "Monto contingencia",
                "Entry (número)",
                "Monto de contingencia. Si tipo producto es crédito/tarjeta debe igualar al investigado.",
                "product_details.csv",
            ),
            (
                "Monto recuperado",
                "Entry (número)",
                "Monto recuperado. No puede ser mayor que el monto investigado.",
                "product_details.csv",
            ),
            (
                "Monto pago deuda",
                "Entry (número)",
                "Monto de pago de deuda. No puede exceder el monto investigado.",
                "product_details.csv",
            ),
            (
                "Botón \"Ir al primer faltante\"",
                "Botón",
                "Navega al reclamo pendiente cuando hay montos > 0.",
                "",
            ),
            (
                "Botón \"Autocompletar analítica\"",
                "Botón",
                "Aplica una analítica preseleccionada al reclamo pendiente.",
                "Catálogo de analíticas",
            ),
        ]
    )

    rows.append(_section_row("Reclamos asociados", columns=len(headers)))
    rows.extend(
        [
            ("Botón \"Añadir reclamo\"", "Botón", "Agrega un reclamo al producto.", ""),
            (
                "ID reclamo",
                "Entry (texto)",
                "Código de reclamo (C + 8 dígitos). Obligatorio si pérdida/falla/contingencia > 0.",
                "claim_details.csv",
            ),
            (
                "Código analítica",
                "Combobox",
                "Código analítica contable de 10 dígitos (43/45/46/56...).",
                "models.analitica_catalog",
            ),
            (
                "Analítica nombre",
                "Combobox",
                "Nombre descriptivo de la analítica contable.",
                "models.analitica_catalog",
            ),
            ("Botón \"Eliminar reclamo\"", "Botón", "Elimina el reclamo del producto.", ""),
        ]
    )

    rows.append(_section_row("Involucramiento de colaboradores", columns=len(headers)))
    rows.extend(
        [
            ("Botón \"Añadir involucrado\"", "Botón", "Agrega un colaborador involucrado.", ""),
            (
                "Colaborador involucrado",
                "Combobox",
                "Selecciona colaborador relacionado con el producto.",
                "Lista de colaboradores",
            ),
            (
                "Monto asignado",
                "Entry (número)",
                "Monto asignado al colaborador (>=0, 12 dígitos, 2 decimales). Requiere ID y viceversa.",
                "No",
            ),
            ("Botón \"Eliminar involucrado\"", "Botón", "Elimina el involucrado.", ""),
        ]
    )

    rows.append(_section_row("Involucramiento de clientes", columns=len(headers)))
    rows.extend(
        [
            ("Botón \"Añadir cliente involucrado\"", "Botón", "Agrega un cliente involucrado.", ""),
            (
                "Cliente involucrado",
                "Combobox",
                "Selecciona cliente relacionado distinto del titular.",
                "Lista de clientes",
            ),
            (
                "Monto asignado",
                "Entry (número)",
                "Monto asignado al cliente involucrado. Requiere ID y viceversa.",
                "No",
            ),
            ("Botón \"Eliminar cliente involucrado\"", "Botón", "Elimina el involucrado.", ""),
            ("Botón \"Eliminar producto\"", "Botón", "Elimina el producto y sus relaciones.", ""),
        ]
    )

    rows.append(_section_row("4. Colaboradores involucrados", columns=len(headers)))
    rows.extend(
        [
            ("Tabla resumen de colaboradores", "Tabla", "Resumen de colaboradores registrados.", ""),
            ("Botón \"Añadir colaborador\"", "Botón", "Crea un colaborador.", ""),
            (
                "ID del colaborador",
                "Entry (texto)",
                "Identificador del Team Member (letra + 5 dígitos).",
                "team_details.csv",
            ),
            ("Nombres", "Entry (texto)", "Nombres del colaborador.", "team_details.csv"),
            ("Apellidos", "Entry (texto)", "Apellidos del colaborador.", "team_details.csv"),
            (
                "Flag",
                "Combobox",
                "Rol del colaborador (Involucrado/Relacionado/etc.).",
                "settings.FLAG_COLABORADOR_LIST",
            ),
            (
                "División",
                "Combobox",
                "División del colaborador (catálogo jerárquico).",
                "TEAM_HIERARCHY_CATALOG",
            ),
            (
                "Área",
                "Combobox",
                "Área dependiente de la división.",
                "TEAM_HIERARCHY_CATALOG",
            ),
            (
                "Servicio",
                "Combobox",
                "Servicio dependiente del área.",
                "TEAM_HIERARCHY_CATALOG",
            ),
            (
                "Puesto",
                "Combobox",
                "Puesto dependiente del servicio.",
                "TEAM_HIERARCHY_CATALOG",
            ),
            (
                "Fecha carta inmediatez (YYYY-MM-DD)",
                "Selector de fecha",
                "Fecha de emisión de carta de inmediatez.",
                "team_details.csv",
            ),
            (
                "Fecha carta renuncia (YYYY-MM-DD)",
                "Selector de fecha",
                "Fecha de renuncia del colaborador.",
                "team_details.csv",
            ),
            (
                "Nombre agencia",
                "Combobox",
                "Nombre de agencia (requerido si División=DCA/Canales y Área contiene 'area comercial').",
                "AGENCY_CATALOG",
            ),
            (
                "Código agencia",
                "Combobox",
                "Código de agencia (6 dígitos) requerido bajo condición de área comercial.",
                "AGENCY_CATALOG",
            ),
            (
                "Tipo de falta",
                "Combobox",
                "Clasificación de falta.",
                "settings.TIPO_FALTA_LIST",
            ),
            (
                "Tipo de sanción",
                "Combobox",
                "Clasificación de sanción.",
                "settings.TIPO_SANCION_LIST",
            ),
            ("Botón \"Eliminar colaborador\"", "Botón", "Elimina el colaborador del caso.", ""),
        ]
    )

    rows = _append_classification(
        rows,
        field_type_index=1,
        source_index=3,
        description_index=2,
        section_label=CLASSIFY_SYSTEM,
    )
    return [headers, *rows]


def _risk_rows() -> list[Sequence[str | None]]:
    headers = ("Campo", "Tipo", "Descripción", "Autocompletado / Fuente", "Clasificación")
    rows: list[Sequence[str | None]] = []
    rows.append(_section_row("Riesgos identificados", columns=len(headers)))
    rows.extend(
        [
            ("Tabla resumen de riesgos", "Tabla", "Listado de riesgos registrados.", ""),
            ("Botón \"Agregar riesgo\"", "Botón", "Crea un riesgo en el caso.", ""),
            (
                "ID de riesgo",
                "Entry (texto)",
                "Identificador del riesgo (catálogo o libre).",
                "risk_details.csv",
            ),
            (
                "Agregar riesgo nuevo",
                "Checkbutton",
                "Activa el modo manual para riesgos no catalogados.",
                "No",
            ),
            (
                "Criticidad",
                "Combobox",
                "Severidad del riesgo (obligatoria en modo catálogo).",
                "settings.CRITICIDAD_LIST / risk_details.csv",
            ),
            ("Líder", "Entry (texto)", "Responsable del riesgo.", "risk_details.csv"),
            (
                "Exposición residual (US$)",
                "Entry (número)",
                "Monto estimado (>=0, 12 dígitos, 2 decimales).",
                "risk_details.csv",
            ),
            (
                "Descripción del riesgo",
                "Entry (texto)",
                "Descripción clara del riesgo.",
                "risk_details.csv",
            ),
            (
                "Planes de acción (IDs separados por ;)",
                "Entry (texto)",
                "Lista de planes asociados sin duplicados.",
                "risk_details.csv",
            ),
            ("Botón \"Eliminar riesgo\"", "Botón", "Elimina el riesgo del caso.", ""),
        ]
    )
    rows = _append_classification(
        rows,
        field_type_index=1,
        source_index=3,
        description_index=2,
        section_label=CLASSIFY_SYSTEM,
    )
    return [headers, *rows]


def _norm_rows() -> list[Sequence[str | None]]:
    headers = ("Campo", "Tipo", "Descripción", "Autocompletado / Fuente", "Clasificación")
    rows: list[Sequence[str | None]] = []
    rows.append(_section_row("Registro de Normas", columns=len(headers)))
    rows.extend(
        [
            ("Tabla resumen de normas", "Tabla", "Listado de normas transgredidas.", ""),
            ("Botón \"Agregar norma\"", "Botón", "Crea una norma en el caso.", ""),
            (
                "ID de norma",
                "Entry (texto)",
                "Formato XXXX.XXX.XX.XX. Permite autopoblado desde catálogo.",
                "norm_details.csv",
            ),
            (
                "Fecha de vigencia (YYYY-MM-DD)",
                "Selector de fecha",
                "Fecha de publicación/vigencia. No puede ser futura.",
                "norm_details.csv",
            ),
            (
                "Descripción de la norma",
                "Entry (texto)",
                "Descripción o título de la norma transgredida.",
                "norm_details.csv",
            ),
            (
                "Acápite/Inciso",
                "Entry (texto)",
                "Referencia del acápite o inciso aplicable.",
                "norm_details.csv",
            ),
            (
                "Detalle de norma",
                "Texto enriquecido",
                "Detalle narrativo del incumplimiento.",
                "norm_details.csv",
            ),
            ("Botón \"Eliminar norma\"", "Botón", "Elimina la norma del caso.", ""),
        ]
    )
    rows = _append_classification(
        rows,
        field_type_index=1,
        source_index=3,
        description_index=2,
        section_label=CLASSIFY_SYSTEM,
    )
    return [headers, *rows]


def _analysis_rows() -> list[Sequence[str | None]]:
    headers = ("Campo", "Tipo", "Descripción", "Autocompletado / Fuente", "Clasificación")
    rows: list[Sequence[str | None]] = []
    rows.append(_section_row("Análisis narrativo", columns=len(headers)))
    rows.extend(
        [
            (
                "Mostrar secciones extendidas del informe",
                "Checkbutton",
                "Habilita el notebook interno con secciones extendidas.",
                "No",
            ),
            ("Antecedentes", "Texto enriquecido", "Contexto del caso.", ""),
            ("Modus operandi", "Texto enriquecido", "Forma de ejecución del fraude.", ""),
            ("Hallazgos principales", "Texto enriquecido", "Hallazgos clave.", ""),
            ("Descargos del colaborador", "Texto enriquecido", "Descargos formales.", ""),
            ("Conclusiones", "Texto enriquecido", "Conclusiones generales.", ""),
            ("Recomendaciones y mejoras", "Texto enriquecido", "Acciones correctivas.", ""),
            (
                "Comentario breve",
                "Texto enriquecido",
                "Resumen sin saltos de línea (máx. 150 caracteres).",
                "",
            ),
            (
                "Comentario amplio",
                "Texto enriquecido",
                "Resumen amplio sin saltos de línea (máx. 750 caracteres).",
                "",
            ),
            (
                "Botón \"Auto-redactar\" (comentarios)",
                "Botón",
                "Genera un resumen automático sin PII.",
                "Modelo LLM",
            ),
        ]
    )
    rows.append(_section_row("Secciones extendidas del informe", columns=len(headers)))
    rows.extend(
        [
            ("Dirigido a", "Entry (texto)", "Destinatario del informe.", ""),
            ("Referencia", "Entry (texto)", "Referencia interna del caso.", ""),
            ("Área de reporte", "Entry (texto)", "Área que emite el informe.", ""),
            (
                "Fecha de reporte (YYYY-MM-DD)",
                "Entry (texto)",
                "Fecha del reporte. Debe ser <= hoy.",
                "",
            ),
            ("Tipología de evento", "Entry (texto)", "Tipología del evento.", ""),
            (
                "Centro de costos (; separados)",
                "Entry (texto)",
                "Centros de costos (numéricos, >=5 dígitos).",
                "",
            ),
            (
                "Procesos impactados",
                "Entry (texto)",
                "Lista de procesos impactados.",
                "",
            ),
            (
                "N° de reclamos",
                "Entry (texto)",
                "Cantidad total de reclamos (numérico).",
                "",
            ),
            (
                "Analítica contable",
                "Combobox",
                "Código/Nombre de analítica contable del catálogo.",
                "models.analitica_catalog",
            ),
            ("Producto (texto opcional)", "Entry (texto)", "Producto objetivo en texto libre.", ""),
            ("Recomendaciones categorizadas - Laboral", "Texto enriquecido", "Lista por ámbito laboral.", ""),
            ("Recomendaciones categorizadas - Operativo", "Texto enriquecido", "Lista por ámbito operativo.", ""),
            ("Recomendaciones categorizadas - Legal", "Texto enriquecido", "Lista por ámbito legal.", ""),
            (
                "Investigador principal - Matrícula/ID",
                "Entry (solo lectura)",
                "Se sincroniza desde Datos generales del caso.",
                "team_details.csv",
            ),
            (
                "Investigador principal - Nombre",
                "Entry (solo lectura)",
                "Autocompletado desde catálogos.",
                "team_details.csv",
            ),
            (
                "Investigador principal - Cargo",
                "Entry (solo lectura)",
                "Autocompletado desde catálogos.",
                "team_details.csv",
            ),
            ("Tabla de operaciones", "Tabla", "Registra operaciones vinculadas.", ""),
            (
                "Operación - N°",
                "Entry (texto)",
                "Número correlativo de la operación.",
                "",
            ),
            (
                "Operación - Fecha aprobación",
                "Entry (texto)",
                "Fecha de aprobación (YYYY-MM-DD).",
                "",
            ),
            (
                "Operación - Cliente/DNI",
                "Entry (texto)",
                "Cliente o documento asociado a la operación.",
                "",
            ),
            (
                "Operación - Ingreso bruto mensual",
                "Entry (texto)",
                "Monto del ingreso bruto mensual.",
                "",
            ),
            (
                "Operación - Empresa empleadora",
                "Entry (texto)",
                "Empresa empleadora asociada.",
                "",
            ),
            (
                "Operación - Vendedor inmueble",
                "Entry (texto)",
                "Vendedor del inmueble.",
                "",
            ),
            (
                "Operación - Vendedor crédito",
                "Entry (texto)",
                "Vendedor del crédito.",
                "",
            ),
            (
                "Operación - Producto",
                "Entry (texto)",
                "Producto asociado a la operación.",
                "",
            ),
            (
                "Operación - Importe desembolsado",
                "Entry (número)",
                "Monto con validación de 2 decimales.",
                "",
            ),
            (
                "Operación - Saldo deudor",
                "Entry (número)",
                "Monto con validación de 2 decimales.",
                "",
            ),
            (
                "Operación - Status BCP",
                "Entry (texto)",
                "Estado del producto en BCP.",
                "",
            ),
            (
                "Operación - Status SBS",
                "Entry (texto)",
                "Estado del producto en SBS.",
                "",
            ),
            ("Botón \"Agregar/Actualizar operación\"", "Botón", "Guarda o actualiza la operación.", ""),
            ("Botón \"Eliminar operación\"", "Botón", "Elimina la operación seleccionada.", ""),
            ("Botón \"Limpiar formulario\"", "Botón", "Limpia los campos de operación.", ""),
            ("Anexos y respaldos", "Tabla", "Control de anexos adjuntos.", ""),
            ("Anexo - Título", "Entry (texto)", "Título del anexo.", ""),
            ("Anexo - Descripción", "Entry (texto)", "Descripción del anexo.", ""),
            ("Botón \"Agregar/Actualizar anexo\"", "Botón", "Guarda el anexo en la tabla.", ""),
            ("Botón \"Eliminar anexo\"", "Botón", "Elimina el anexo seleccionado.", ""),
            ("Botón \"Limpiar\" (anexo)", "Botón", "Limpia el formulario de anexos.", ""),
        ]
    )
    rows = _append_classification(
        rows,
        field_type_index=1,
        source_index=3,
        description_index=2,
        section_label=CLASSIFY_SYSTEM,
    )
    return [headers, *rows]


def _actions_rows() -> list[Sequence[str | None]]:
    headers = ("Campo", "Tipo", "Descripción", "Autocompletado / Fuente", "Clasificación")
    rows: list[Sequence[str | None]] = []
    rows.append(_section_row("Acciones", columns=len(headers)))
    rows.extend(
        [
            (
                "Sonido de confirmación",
                "Checkbutton",
                "Activa/desactiva sonido tras validaciones y exportes.",
                "Preferencias usuario",
            ),
            (
                "Botón conmutar tema",
                "Botón",
                "Cambia el tema oscuro/claro.",
                "ThemeManager",
            ),
            ("Catálogos de detalle", "Sección", "Controles para carga de catálogos.", ""),
            ("Estado/ayuda de catálogos", "Label", "Estado de carga de catálogos.", ""),
            ("Botón \"Cargar catálogos\"", "Botón", "Carga catálogos desde CSV.", ""),
            ("Botón \"Iniciar sin catálogos\"", "Botón", "Continúa sin catálogos.", ""),
            ("Barra de progreso de catálogos", "Progressbar", "Indicador de carga.", ""),
            ("Importar datos masivos (CSV)", "Sección", "Carga de archivos masivos.", ""),
            ("Botón \"Cargar clientes\"", "Botón", "Importa clientes masivos.", "clientes_masivos.csv"),
            ("Botón \"Cargar colaboradores\"", "Botón", "Importa colaboradores masivos.", "colaboradores_masivos.csv"),
            ("Botón \"Cargar productos\"", "Botón", "Importa productos masivos.", "productos_masivos.csv"),
            ("Botón \"Cargar combinado\"", "Botón", "Importa clientes/productos/colaboradores.", "datos_combinados_masivos.csv"),
            ("Botón \"Cargar riesgos\"", "Botón", "Importa riesgos masivos.", "riesgos_masivos.csv"),
            ("Botón \"Cargar normas\"", "Botón", "Importa normas masivas.", "normas_masivas.csv"),
            ("Botón \"Cargar reclamos\"", "Botón", "Importa reclamos masivos.", "reclamos_masivos.csv"),
            ("Estado de importación", "Label", "Estado de la importación masiva.", ""),
            ("Barra de progreso de importación", "Progressbar", "Indicador de importación.", ""),
            ("Guardar, cargar y reportes", "Sección", "Acciones de guardado y reportes.", ""),
            ("Botón \"Guardar ahora\"", "Botón", "Valida y guarda exportes.", ""),
            ("Botón \"Cargar archivo…\"", "Botón", "Carga un respaldo JSON.", ""),
            ("Botón \"Recuperar último autosave\"", "Botón", "Recupera último autosave.", ""),
            ("Botón \"Historial de recuperación\"", "Botón", "Abre historial de versiones.", ""),
            ("Botón \"Generar informe (.md)\"", "Botón", "Genera informe Markdown.", ""),
            ("Botón \"Generar Word (.docx)\"", "Botón", "Genera informe Word.", ""),
            ("Botón \"Generar alerta temprana (.pptx)\"", "Botón", "Genera alerta temprana PPT.", ""),
            ("Botón \"Generar carta de inmediatez\"", "Botón", "Genera carta de inmediatez.", ""),
            ("Botón \"Borrar todos los datos\"", "Botón", "Limpia el formulario.", ""),
            (
                "Codificación de exportación",
                "Combobox",
                "Selecciona codificación para CSV (UTF-8 recomendado).",
                "Preferencias usuario",
            ),
        ]
    )
    rows = _append_classification(
        rows,
        field_type_index=1,
        source_index=3,
        description_index=2,
        section_label=CLASSIFY_SYSTEM,
    )
    return [headers, *rows]


def _summary_rows() -> list[Sequence[str | None]]:
    headers = ("Sección/Tabla", "Tipo", "Descripción", "Autocompletado / Fuente", "Clasificación")
    rows: list[Sequence[str | None]] = []
    rows.extend(
        [
            (
                "Label introductorio",
                "Label",
                "Introduce la sección de resumen con actualización automática.",
                "",
            ),
            (
                "Clientes registrados",
                "Tabla",
                "Columnas: ID Cliente, Nombres, Apellidos, Tipo ID, Flag, Teléfonos, Correos, Direcciones, Accionado.",
                "",
            ),
            (
                "Colaboradores involucrados",
                "Tabla",
                "Columnas: ID, Nombres, Apellidos, Flag, División, Área, Servicio, Puesto, Fechas, Agencia, Falta, Sanción.",
                "",
            ),
            (
                "Asignaciones de involucrados",
                "Tabla",
                "Columnas: Producto, Tipo, Colaborador/Cliente involucrado, Monto asignado.",
                "",
            ),
            (
                "Productos investigados",
                "Tabla",
                "Columnas: ID Producto, Cliente, Tipo, Taxonomía, Fechas, Montos, Reclamos.",
                "",
            ),
            (
                "Riesgos registrados",
                "Tabla",
                "Columnas: ID Riesgo, Líder, Descripción, Criticidad, Exposición, Planes.",
                "",
            ),
            (
                "Reclamos asociados",
                "Tabla",
                "Columnas: ID Reclamo, ID Caso, ID Producto, Analítica, Código analítica.",
                "",
            ),
            (
                "Normas transgredidas",
                "Tabla",
                "Columnas: ID Norma, ID Caso, Descripción, Vigencia, Acápite/Inciso, Detalle.",
                "",
            ),
        ]
    )
    rows = _append_classification(
        rows,
        field_type_index=1,
        source_index=3,
        description_index=2,
        section_label=CLASSIFY_SYSTEM,
    )
    return [headers, *rows]


def _export_structure_rows() -> list[Sequence[str | None]]:
    headers = ("Archivo", "Columna", "Tipo", "Descripción/Notas", "Clasificación")
    rows: list[Sequence[str | None]] = []
    for file_name, fields in EXPORT_HEADERS.items():
        rows.append(_section_row(file_name, columns=len(headers)))
        for field in fields:
            if field.startswith("fecha"):
                field_type = "date"
            elif field.startswith("monto") or field in {"exposicion_residual"}:
                field_type = "decimal"
            else:
                field_type = "string"
            description = EXPORT_FIELD_DESCRIPTIONS.get(file_name, {}).get(field, "")
            rows.append((file_name, field, field_type, description, CLASSIFY_EXPORT))
    rows.append(_section_row("eventos.csv", columns=len(headers)))
    rows.append(("eventos.csv", "(ver hoja Eventos_CSV)", "string", "El esquema canónico se detalla aparte.", CLASSIFY_EXPORT))
    rows.append(_section_row("eventos_lhcl.csv", columns=len(headers)))
    rows.append(
        (
            "eventos_lhcl.csv",
            "(igual que eventos.csv)",
            "string",
            "Export adicional con el mismo header canónico.",
            CLASSIFY_EXPORT,
        )
    )
    rows.append(_section_row("llave_tecnica.csv", columns=len(headers)))
    rows.append(
        (
            "llave_tecnica.csv",
            "(ver build_llave_tecnica_rows)",
            "string",
            "Combina caso + producto + involucrados + reclamo.",
            CLASSIFY_EXPORT,
        )
    )
    rows.append(_section_row("logs.csv", columns=len(headers)))
    for field in LOG_FIELDNAMES:
        rows.append(("logs.csv", field, "string", "Bitácora de eventos y validaciones.", CLASSIFY_EXPORT))
    return [headers, *rows]


def _mapping_export_rows() -> list[Sequence[str | None]]:
    headers = ("Archivo", "Columna", "Origen (pestaña/campo)", "Transformación / Notas", "Clasificación")
    rows: list[Sequence[str | None]] = []

    case_map = {
        "id_caso": "Caso y participantes > Número de caso",
        "tipo_informe": "Caso y participantes > Tipo de informe",
        "categoria1": "Caso y participantes > Categoría nivel 1",
        "categoria2": "Caso y participantes > Categoría nivel 2",
        "modalidad": "Caso y participantes > Modalidad",
        "canal": "Caso y participantes > Canal",
        "proceso": "Caso y participantes > Proceso impactado",
        "fecha_de_ocurrencia": "Caso y participantes > Ocurrencia",
        "fecha_de_descubrimiento": "Caso y participantes > Descubrimiento",
        "centro_costos": "Caso y participantes > Centro de costos",
        "matricula_investigador": "Caso y participantes > Matrícula investigador",
        "investigador_nombre": "Caso y participantes > Nombre y apellidos",
        "investigador_cargo": "Caso y participantes > Puesto",
    }
    for field in EXPORT_HEADERS["casos.csv"]:
        rows.append(("casos.csv", field, case_map.get(field, ""), "", CLASSIFY_MAPPING))

    client_map = {
        "id_cliente": "Clientes > ID del cliente",
        "id_caso": "Caso y participantes > Número de caso",
        "nombres": "Clientes > Nombres",
        "apellidos": "Clientes > Apellidos",
        "tipo_id": "Clientes > Tipo de ID",
        "flag": "Clientes > Flag",
        "telefonos": "Clientes > Teléfonos",
        "correos": "Clientes > Correos",
        "direcciones": "Clientes > Direcciones",
        "accionado": "Clientes > Accionado",
    }
    for field in EXPORT_HEADERS["clientes.csv"]:
        rows.append(("clientes.csv", field, client_map.get(field, ""), "", CLASSIFY_MAPPING))

    team_map = {
        "id_colaborador": "Colaboradores > ID del colaborador",
        "id_caso": "Caso y participantes > Número de caso",
        "flag": "Colaboradores > Flag",
        "nombres": "Colaboradores > Nombres",
        "apellidos": "Colaboradores > Apellidos",
        "division": "Colaboradores > División",
        "area": "Colaboradores > Área",
        "servicio": "Colaboradores > Servicio",
        "puesto": "Colaboradores > Puesto",
        "fecha_carta_inmediatez": "Colaboradores > Fecha carta inmediatez",
        "fecha_carta_renuncia": "Colaboradores > Fecha carta renuncia",
        "nombre_agencia": "Colaboradores > Nombre agencia",
        "codigo_agencia": "Colaboradores > Código agencia",
        "tipo_falta": "Colaboradores > Tipo de falta",
        "tipo_sancion": "Colaboradores > Tipo de sanción",
    }
    for field in EXPORT_HEADERS["colaboradores.csv"]:
        rows.append(("colaboradores.csv", field, team_map.get(field, ""), "", CLASSIFY_MAPPING))

    product_map = {
        "id_producto": "Productos > ID del producto",
        "id_caso": "Caso y participantes > Número de caso",
        "id_cliente": "Productos > Cliente",
        "categoria1": "Productos > Categoría 1",
        "categoria2": "Productos > Categoría 2",
        "modalidad": "Productos > Modalidad",
        "canal": "Productos > Canal",
        "proceso": "Productos > Proceso",
        "fecha_ocurrencia": "Productos > Fecha de ocurrencia",
        "fecha_descubrimiento": "Productos > Fecha de descubrimiento",
        "monto_investigado": "Productos > Monto investigado",
        "tipo_moneda": "Productos > Moneda",
        "monto_perdida_fraude": "Productos > Monto pérdida fraude",
        "monto_falla_procesos": "Productos > Monto falla procesos",
        "monto_contingencia": "Productos > Monto contingencia",
        "monto_recuperado": "Productos > Monto recuperado",
        "monto_pago_deuda": "Productos > Monto pago deuda",
        "tipo_producto": "Productos > Tipo de producto",
    }
    product_notes = {
        "categoria1": "Override por producto; si está vacío se usa la taxonomía del caso.",
        "categoria2": "Override por producto; si está vacío se usa la taxonomía del caso.",
        "modalidad": "Override por producto; si está vacío se usa la taxonomía del caso.",
        "canal": "Override por producto; si está vacío se usa el canal del caso.",
        "proceso": "Override por producto; si está vacío se usa el proceso del caso.",
    }
    for field in EXPORT_HEADERS["productos.csv"]:
        rows.append(
            ("productos.csv", field, product_map.get(field, ""), product_notes.get(field, ""), CLASSIFY_MAPPING)
        )

    claim_map = {
        "id_reclamo": "Productos > Reclamos asociados > ID reclamo",
        "id_caso": "Caso y participantes > Número de caso",
        "id_producto": "Productos > ID del producto",
        "nombre_analitica": "Productos > Reclamos asociados > Analítica nombre",
        "codigo_analitica": "Productos > Reclamos asociados > Código analítica",
    }
    for field in EXPORT_HEADERS["producto_reclamo.csv"]:
        rows.append(("producto_reclamo.csv", field, claim_map.get(field, ""), "", CLASSIFY_MAPPING))

    inv_map = {
        "id_producto": "Productos > ID del producto",
        "id_caso": "Caso y participantes > Número de caso",
        "tipo_involucrado": "Productos > Involucramientos",
        "id_colaborador": "Productos > Involucramiento colaboradores",
        "id_cliente_involucrado": "Productos > Involucramiento clientes",
        "monto_asignado": "Productos > Monto asignado",
    }
    inv_notes = {
        "tipo_involucrado": "Valores esperados: cliente/colaborador.",
        "id_colaborador": "Se llena solo si el tipo de involucrado es colaborador.",
        "id_cliente_involucrado": "Se llena solo si el tipo de involucrado es cliente.",
    }
    for field in EXPORT_HEADERS["involucramiento.csv"]:
        rows.append(("involucramiento.csv", field, inv_map.get(field, ""), inv_notes.get(field, ""), CLASSIFY_MAPPING))

    risk_map = {
        "id_riesgo": "Riesgos > ID de riesgo",
        "id_caso": "Caso y participantes > Número de caso",
        "lider": "Riesgos > Líder",
        "descripcion": "Riesgos > Descripción",
        "criticidad": "Riesgos > Criticidad",
        "exposicion_residual": "Riesgos > Exposición residual",
        "planes_accion": "Riesgos > Planes de acción",
    }
    for field in EXPORT_HEADERS["detalles_riesgo.csv"]:
        rows.append(("detalles_riesgo.csv", field, risk_map.get(field, ""), "", CLASSIFY_MAPPING))

    norm_map = {
        "id_norma": "Normas > ID de norma",
        "id_caso": "Caso y participantes > Número de caso",
        "descripcion": "Normas > Descripción",
        "fecha_vigencia": "Normas > Fecha vigencia",
        "acapite_inciso": "Normas > Acápite/Inciso",
        "detalle_norma": "Normas > Detalle de norma",
    }
    for field in EXPORT_HEADERS["detalles_norma.csv"]:
        rows.append(("detalles_norma.csv", field, norm_map.get(field, ""), "", CLASSIFY_MAPPING))

    analysis_map = {
        "id_caso": "Caso y participantes > Número de caso",
        "antecedentes": "Análisis > Antecedentes",
        "modus_operandi": "Análisis > Modus operandi",
        "hallazgos": "Análisis > Hallazgos principales",
        "descargos": "Análisis > Descargos",
        "conclusiones": "Análisis > Conclusiones",
        "recomendaciones": "Análisis > Recomendaciones",
        "comentario_breve": "Análisis > Comentario breve",
        "comentario_amplio": "Análisis > Comentario amplio",
    }
    for field in EXPORT_HEADERS["analisis.csv"]:
        rows.append(("analisis.csv", field, analysis_map.get(field, ""), "", CLASSIFY_MAPPING))

    return [headers, *rows]


def _eventos_rows() -> list[Sequence[str | None]]:
    headers = ("Columna", "Descripción", "Origen (pestaña/campo)", "Transformación / Notas", "Clasificación")
    rows: list[Sequence[str | None]] = []
    alias_map = {
        "case_id": "id_caso",
        "id_caso": "case_id",
        "product_id": "id_producto",
        "id_producto": "product_id",
        "client_id_involucrado": "id_cliente_involucrado",
        "id_cliente_involucrado": "client_id_involucrado",
        "matricula_colaborador_involucrado": "id_colaborador",
        "id_colaborador": "matricula_colaborador_involucrado",
        "fecha_ocurrencia_caso": "fecha_ocurrencia",
        "fecha_descubrimiento_caso": "fecha_descubrimiento",
    }
    origin_map = {
        "case_id": "Caso > Número de caso",
        "id_caso": "Caso > Número de caso",
        "tipo_informe": "Caso > Tipo de informe",
        "categoria_1": "Productos (override) / Caso > Taxonomía",
        "categoria_2": "Productos (override) / Caso > Taxonomía",
        "modalidad": "Productos (override) / Caso > Taxonomía",
        "categoria1": "Caso > Categoría nivel 1",
        "categoria2": "Caso > Categoría nivel 2",
        "canal": "Productos (override) / Caso > Canal",
        "proceso_impactado": "Productos (override) / Caso > Proceso",
        "proceso": "Caso > Proceso impactado",
        "fecha_de_ocurrencia": "Caso > Fecha de ocurrencia",
        "fecha_de_descubrimiento": "Caso > Fecha de descubrimiento",
        "fecha_ocurrencia_caso": "Caso > Fecha de ocurrencia",
        "fecha_descubrimiento_caso": "Caso > Fecha de descubrimiento",
        "centro_costos": "Caso > Centro de costos",
        "matricula_investigador": "Caso > Matrícula investigador",
        "investigador_nombre": "Caso > Nombre investigador",
        "investigador_cargo": "Caso > Cargo investigador",
        "tipo_de_producto": "Productos > Tipo de producto",
        "tipo_producto": "Productos > Tipo de producto",
        "product_id": "Productos > ID del producto",
        "id_producto": "Productos > ID del producto",
        "tipo_moneda": "Productos > Moneda",
        "monto_investigado": "Productos > Monto investigado",
        "monto_perdida_fraude": "Productos > Monto pérdida fraude",
        "monto_falla_procesos": "Productos > Monto falla procesos",
        "monto_contingencia": "Productos > Monto contingencia",
        "monto_recuperado": "Productos > Monto recuperado",
        "monto_pago_deuda": "Productos > Monto pago deuda",
        "fecha_ocurrencia": "Productos > Fecha ocurrencia (fallback caso)",
        "fecha_descubrimiento": "Productos > Fecha descubrimiento (fallback caso)",
        "comentario_breve": "Análisis > Comentario breve",
        "comentario_amplio": "Análisis > Comentario amplio",
        "id_reclamo": "Productos > Reclamos asociados",
        "nombre_analitica": "Productos > Reclamos asociados",
        "codigo_analitica": "Productos > Reclamos asociados",
        "tipo_involucrado": "Productos > Involucramientos",
        "id_colaborador": "Productos > Involucramientos",
        "id_cliente_involucrado": "Productos > Involucramientos",
        "monto_asignado": "Productos > Involucramientos",
    }
    client_involved_origin = {
        "tipo_id_cliente_involucrado": "Clientes (involucrado) > Tipo ID",
        "client_id_involucrado": "Clientes (involucrado) > ID",
        "flag_cliente_involucrado": "Clientes (involucrado) > Flag",
        "nombres_cliente_involucrado": "Clientes (involucrado) > Nombres",
        "apellidos_cliente_involucrado": "Clientes (involucrado) > Apellidos",
    }
    collaborator_involved_origin = {
        "matricula_colaborador_involucrado": "Colaboradores (involucrado) > ID",
        "apellido_paterno_involucrado": "Colaboradores (involucrado) > Apellidos",
        "apellido_materno_involucrado": "Colaboradores (involucrado) > Apellidos",
        "nombres_involucrado": "Colaboradores (involucrado) > Nombres",
        "division": "Colaboradores (involucrado) > División",
        "area": "Colaboradores (involucrado) > Área",
        "servicio": "Colaboradores (involucrado) > Servicio",
        "nombre_agencia": "Colaboradores (involucrado) > Nombre agencia",
        "codigo_agencia": "Colaboradores (involucrado) > Código agencia",
        "puesto": "Colaboradores (involucrado) > Puesto",
        "fecha_cese": "Colaboradores (involucrado) > Fecha carta renuncia",
        "tipo_de_falta": "Colaboradores (involucrado) > Tipo de falta",
        "tipo_sancion": "Colaboradores (involucrado) > Tipo de sanción",
    }
    client_owner_origin = {
        "telefonos_cliente_relacionado": "Clientes (titular) > Teléfonos",
        "correos_cliente_relacionado": "Clientes (titular) > Correos",
        "direcciones_cliente_relacionado": "Clientes (titular) > Direcciones",
        "accionado_cliente_relacionado": "Clientes (titular) > Accionado",
        "cliente_nombres": "Clientes (titular) > Nombres",
        "cliente_apellidos": "Clientes (titular) > Apellidos",
        "cliente_tipo_id": "Clientes (titular) > Tipo ID",
        "cliente_flag": "Clientes (titular) > Flag",
        "cliente_telefonos": "Clientes (titular) > Teléfonos",
        "cliente_correos": "Clientes (titular) > Correos",
        "cliente_direcciones": "Clientes (titular) > Direcciones",
        "cliente_accionado": "Clientes (titular) > Accionado",
        "id_cliente": "Clientes (titular) > ID",
    }
    collaborator_origin = {
        "colaborador_flag": "Colaboradores > Flag",
        "colaborador_nombres": "Colaboradores > Nombres",
        "colaborador_apellidos": "Colaboradores > Apellidos",
        "colaborador_division": "Colaboradores > División",
        "colaborador_area": "Colaboradores > Área",
        "colaborador_servicio": "Colaboradores > Servicio",
        "colaborador_puesto": "Colaboradores > Puesto",
        "colaborador_fecha_carta_inmediatez": "Colaboradores > Fecha carta inmediatez",
        "colaborador_fecha_carta_renuncia": "Colaboradores > Fecha carta renuncia",
        "colaborador_nombre_agencia": "Colaboradores > Nombre agencia",
        "colaborador_codigo_agencia": "Colaboradores > Código agencia",
        "colaborador_tipo_falta": "Colaboradores > Tipo de falta",
        "colaborador_tipo_sancion": "Colaboradores > Tipo de sanción",
    }
    monto_detail_origin = {
        "monto_fraude_interno_soles": "Productos > Monto pérdida fraude",
        "monto_falla_en_proceso_soles": "Productos > Monto falla procesos",
        "monto_contingencia_soles": "Productos > Monto contingencia",
        "monto_recuperado_soles": "Productos > Monto recuperado",
        "monto_pagado_soles": "Productos > Monto pago deuda",
    }
    placeholder_notes = {
        "cod_operation": "Siempre se exporta como <SIN_DATO>.",
        "apellido_materno_involucrado": "No se captura; se exporta como <SIN_DATO>.",
        "monto_fraude_externo_soles": "Actualmente se exporta como <SIN_DATO>.",
    }
    canonical_fields = set(EVENTOS_HEADER_CANONICO_START)
    for field in EVENTOS_HEADER_CANONICO:
        if field in canonical_fields:
            description = "Campo canónico de eventos."
        else:
            description = "Campo legado incluido por compatibilidad histórica."
        origin = (
            origin_map.get(field)
            or client_involved_origin.get(field)
            or collaborator_involved_origin.get(field)
            or client_owner_origin.get(field)
            or collaborator_origin.get(field)
            or monto_detail_origin.get(field)
            or ""
        )
        notes = placeholder_notes.get(field, "")
        if not origin and field.startswith("monto_"):
            origin = "Productos > Montos"
        if field == "cod_operation":
            origin = "No aplica (legacy)"
        if field.endswith("_dolares"):
            notes = "Actualmente se llena con <SIN_DATO>."
        if field in client_involved_origin or field in {"client_id_involucrado", "id_cliente_involucrado"}:
            notes = f"{notes} Solo aplica si el involucrado es cliente.".strip()
        if field in collaborator_involved_origin:
            notes = f"{notes} Solo aplica si el involucrado es colaborador.".strip()
        if field in alias_map:
            alias_note = f"Alias/compatibilidad con {alias_map[field]}."
            notes = f"{notes} {alias_note}".strip()
        rows.append((field, description, origin, notes, CLASSIFY_EXPORT))
    return [headers, *rows]


def _logs_rows() -> list[Sequence[str | None]]:
    headers = ("Campo", "Descripción", "Origen / Uso", "Clasificación")
    rows = [
        ("timestamp", "Fecha y hora del evento registrado.", "Generado automáticamente.", CLASSIFY_SYSTEM),
        ("tipo", "Tipo general del evento (validacion, navegacion, etc.).", "log_event", CLASSIFY_SYSTEM),
        ("subtipo", "Subtipo o etiqueta específica.", "log_event", CLASSIFY_SYSTEM),
        ("widget_id", "Identificador del widget asociado.", "Widget registry", CLASSIFY_SYSTEM),
        ("coords", "Coordenadas del cursor.", "log_event", CLASSIFY_SYSTEM),
        ("mensaje", "Mensaje descriptivo del evento.", "Validadores/UI", CLASSIFY_SYSTEM),
        ("old_value", "Valor previo en validaciones.", "FieldValidator", CLASSIFY_SYSTEM),
        ("new_value", "Valor nuevo en validaciones.", "FieldValidator", CLASSIFY_SYSTEM),
        ("action_result", "Resultado de la acción (ok/error).", "FieldValidator", CLASSIFY_SYSTEM),
    ]
    return [headers, *rows]


def _carta_rows() -> list[Sequence[str | None]]:
    headers = ("Campo CSV/Placeholder", "Descripción", "Origen (pestaña/campo)", "Transformación / Notas", "Clasificación")
    rows = [
        ("numero_caso", "ID del caso.", "Caso > Número de caso", "Valor directo", CLASSIFY_REPORT),
        ("fecha_generacion", "Fecha de generación de carta.", "Sistema", "YYYY-MM-DD", CLASSIFY_REPORT),
        ("mes", "Mes de generación.", "Sistema", "Formato MMMM", CLASSIFY_REPORT),
        ("investigador_principal", "Nombre investigador principal.", "Caso > Nombre investigador", "", CLASSIFY_REPORT),
        ("matricula_investigador", "Matrícula del investigador.", "Caso > Matrícula investigador", "", CLASSIFY_REPORT),
        ("matricula_team_member", "ID del colaborador.", "Colaboradores > ID", "Normalizado a mayúsculas", CLASSIFY_REPORT),
        ("Tipo", "Tipo de sede (Agencia/Sede).", "Colaboradores > División", "Agencia si división contiene 'comercial' o 'DCC'", CLASSIFY_REPORT),
        ("codigo_agencia", "Código de agencia.", "Colaboradores > Código agencia", "", CLASSIFY_REPORT),
        ("agencia", "Nombre de agencia.", "Colaboradores > Nombre agencia", "", CLASSIFY_REPORT),
        ("Numero_de_Carta", "Número correlativo de carta.", "Sistema", "Formato 000-AAAA", CLASSIFY_REPORT),
        ("Tipo_entrevista", "Tipo de entrevista.", "Colaboradores > Flag", "Involucrado o Informativo", CLASSIFY_REPORT),
        ("FECHA", "Fecha de generación en plantilla.", "Sistema", "YYYY-MM-DD", CLASSIFY_REPORT),
        ("FECHA_LARGA", "Fecha larga en plantilla DOCX.", "Sistema", "Formato largo (e.g. 12 de marzo de 2025)", CLASSIFY_REPORT),
        ("NOMBRE_COMPLETO", "Nombre del colaborador.", "Colaboradores > Nombres/Apellidos", "Usa matrícula si falta nombre", CLASSIFY_REPORT),
        ("MATRICULA", "Matrícula del colaborador.", "Colaboradores > ID", "", CLASSIFY_REPORT),
        ("APELLIDOS", "Apellidos del colaborador.", "Colaboradores > Apellidos", "", CLASSIFY_REPORT),
        ("AREA", "Área del colaborador.", "Colaboradores > Área", "", CLASSIFY_REPORT),
        ("NUMERO_CARTA", "Número de carta.", "Sistema", "Alias de Numero_de_Carta", CLASSIFY_REPORT),
        ("NUMERO_CASO", "Número de caso.", "Caso > Número de caso", "", CLASSIFY_REPORT),
        ("COLABORADOR", "Nombre visible del colaborador.", "Colaboradores > Nombres/Apellidos", "", CLASSIFY_REPORT),
        ("PUESTO", "Puesto del colaborador.", "Colaboradores > Puesto", "", CLASSIFY_REPORT),
        ("AGENCIA", "Agencia del colaborador.", "Colaboradores > Nombre agencia", "", CLASSIFY_REPORT),
        ("INVESTIGADOR", "Investigador principal.", "Caso > Nombre investigador", "", CLASSIFY_REPORT),
    ]
    return [headers, *rows]


def _gerencia_rows() -> list[Sequence[str | None]]:
    headers = ("Sección", "Descripción", "Campos y Transformaciones", "Clasificación")
    rows = [
        (
            "Cabecera",
            "Datos generales del caso.",
            "Tipo de informe, número de caso, categoría, canal, proceso, fechas.",
            CLASSIFY_REPORT,
        ),
        (
            "Antecedentes",
            "Narrativa principal.",
            "Texto de Antecedentes.",
            CLASSIFY_REPORT,
        ),
        (
            "Colaboradores involucrados",
            "Resumen de colaboradores.",
            "Datos de pestaña Colaboradores + fechas de cartas.",
            CLASSIFY_REPORT,
        ),
        (
            "Clientes involucrados",
            "Resumen de clientes.",
            "Datos de pestaña Clientes (tipo ID, nombres, flag, contacto).",
            CLASSIFY_REPORT,
        ),
        (
            "Productos",
            "Resumen de productos investigados.",
            "Montos, fechas y reclamos asociados.",
            CLASSIFY_REPORT,
        ),
        (
            "Riesgos Potenciales",
            "Riesgos registrados.",
            "ID riesgo, líder, descripción, criticidad, exposición.",
            CLASSIFY_REPORT,
        ),
        (
            "Normas",
            "Normas transgredidas.",
            "ID norma, descripción, vigencia, acápite, detalle.",
            CLASSIFY_REPORT,
        ),
        (
            "Análisis y Hallazgos",
            "Narrativa consolidada.",
            "Modus operandi + Hallazgos principales.",
            CLASSIFY_REPORT,
        ),
        (
            "Descargos y Testimonios",
            "Sección de descargos.",
            "Descargos del colaborador.",
            CLASSIFY_REPORT,
        ),
        (
            "Conclusiones y Recomendaciones",
            "Cierre del informe.",
            "Conclusiones + Recomendaciones.",
            CLASSIFY_REPORT,
        ),
    ]
    return [headers, *rows]


def _alerta_rows() -> list[Sequence[str | None]]:
    headers = ("Sección de PPT", "Descripción", "Campos / Datos requeridos", "Clasificación")
    rows = [
        (
            "Masthead",
            "Carátula del caso.",
            "ID caso, investigador, referencia/temática (encabezado), categoría, canal.",
            CLASSIFY_REPORT,
        ),
        (
            "Resumen ejecutivo",
            "Mensaje clave con soporte y evidencias.",
            "ID caso + referencia + montos agregados + hallazgos/resumen ejecutivo.",
            CLASSIFY_REPORT,
        ),
        (
            "Resumen",
            "Resumen numérico.",
            "Resumen ejecutivo/Conclusiones + montos agregados.",
            CLASSIFY_REPORT,
        ),
        (
            "Cronología",
            "Fechas clave del caso.",
            "Hallazgos (análisis) + operaciones; fallback a fechas de ocurrencia/descubrimiento.",
            CLASSIFY_REPORT,
        ),
        (
            "Análisis",
            "Narrativa resumida.",
            "Texto consolidado de análisis y hallazgos.",
            CLASSIFY_REPORT,
        ),
        (
            "Riesgos identificados",
            "Listado de riesgos.",
            "ID riesgo, descripción, criticidad.",
            CLASSIFY_REPORT,
        ),
        (
            "Acciones",
            "Acciones/recomendaciones.",
            "Recomendaciones + operaciones vinculadas.",
            CLASSIFY_REPORT,
        ),
        (
            "Responsables",
            "Investigador y responsables.",
            "Investigador principal + colaboradores con flag y área.",
            CLASSIFY_REPORT,
        ),
    ]
    return [headers, *rows]


def _resumen_ejecutivo_rows() -> list[Sequence[str | None]]:
    headers = ("Sección", "Descripción", "Campos / Datos requeridos", "Clasificación")
    rows = [
        (
            "Mensaje clave",
            "Síntesis principal (piramidal).",
            "ID caso, categoría/modalidad, canal/proceso, montos agregados.",
            CLASSIFY_REPORT,
        ),
        (
            "Puntos de soporte (3-5)",
            "Hallazgos, riesgos y acciones relevantes.",
            "Hallazgos/conclusiones, riesgos, recomendaciones/operaciones, responsables.",
            CLASSIFY_REPORT,
        ),
        (
            "Evidencia / trazabilidad",
            "Métricas y referencias.",
            "Conteos (productos/clientes/colaboradores), fechas clave, dirigido a, área reporte.",
            CLASSIFY_REPORT,
        ),
    ]
    return [headers, *rows]


def _validation_panel_rows() -> list[Sequence[str | None]]:
    headers = ("Campo / Validación", "Descripción de la regla", "Mensaje de error", "Fuente", "Clasificación")
    rows = [
        (
            "Número de caso",
            "Formato AAAA-NNNN.",
            "El número de caso debe seguir el formato AAAA-NNNN.",
            "validators.validate_case_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "ID Proceso",
            "Formato BPID-XXXXXX o BPID-RNF-XXXXXX.",
            "El ID de proceso debe seguir el formato BPID-XXXXXX o BPID-RNF-XXXXXX.",
            "validators.validate_process_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "Fechas de caso",
            "Formato YYYY-MM-DD; ocurrencia < descubrimiento; ambas <= hoy.",
            "La fecha de ocurrencia debe ser anterior a la de descubrimiento.",
            "validators.validate_date_text",
            CLASSIFY_VALIDATION,
        ),
        (
            "Centro de costos",
            "Valores numéricos separados por ';' con mínimo 5 dígitos.",
            "Cada centro de costos debe ser numérico y tener al menos 5 dígitos.",
            "app._validate_cost_centers",
            CLASSIFY_VALIDATION,
        ),
        (
            "Fechas de producto",
            "Formato YYYY-MM-DD; ocurrencia < descubrimiento; ambas <= hoy.",
            "Las fechas del producto no pueden estar en el futuro.",
            "validators.validate_product_dates",
            CLASSIFY_VALIDATION,
        ),
        (
            "Fecha de vigencia (norma)",
            "Formato YYYY-MM-DD; no puede ser futura.",
            "La fecha de vigencia no puede estar en el futuro.",
            "validators.validate_date_text",
            CLASSIFY_VALIDATION,
        ),
        (
            "Fecha de vigencia (norma)",
            "Formato YYYY-MM-DD; no puede ser futura.",
            "La fecha de vigencia no puede estar en el futuro.",
            "validators.validate_date_text",
        ),
        (
            "Montos",
            ">=0, 12 dígitos, 2 decimales. Investigado = pérdida+falla+contingencia+recuperado.",
            "La suma de las cuatro partidas debe ser igual al monto investigado.",
            "ui/frames/products.py",
            CLASSIFY_VALIDATION,
        ),
        (
            "Monto pago deuda",
            "Debe ser <= monto investigado.",
            "El pago de deuda no puede ser mayor al monto investigado.",
            "ui/frames/products.py",
            CLASSIFY_VALIDATION,
        ),
        (
            "Monto contingencia (crédito/tarjeta)",
            "Debe igualar monto investigado si tipo producto es crédito o tarjeta.",
            "El monto de contingencia debe ser igual al monto investigado para créditos o tarjetas.",
            "ui/frames/products.py",
            CLASSIFY_VALIDATION,
        ),
        (
            "Correo electrónico",
            "Formato válido para cada correo separado por ';'.",
            "El campo contiene un correo inválido: <correo>.",
            "validators.validate_email_list",
            CLASSIFY_VALIDATION,
        ),
        (
            "Teléfono",
            "Formato +? y 6-15 dígitos, separados por ';'.",
            "El campo contiene un teléfono inválido: <tel>.",
            "validators.validate_phone_list",
            CLASSIFY_VALIDATION,
        ),
        (
            "Accionado (cliente)",
            "Debe seleccionar al menos una opción.",
            "Debe seleccionar al menos una opción en Accionado.",
            "validators.validate_multi_selection",
            CLASSIFY_VALIDATION,
        ),
        (
            "ID de reclamo",
            "Formato C########.",
            "El ID de reclamo debe tener el formato CXXXXXXXX.",
            "validators.validate_reclamo_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "Código de analítica",
            "10 dígitos, inicia 43/45/46/56.",
            "El código de analítica debe tener 10 dígitos y comenzar con 43/45/46/56.",
            "validators.validate_codigo_analitica",
            CLASSIFY_VALIDATION,
        ),
        (
            "ID de norma",
            "Formato XXXX.XXX.XX.XX.",
            "El ID de norma debe seguir el formato XXXX.XXX.XX.XX.",
            "validators.validate_norm_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "ID de riesgo",
            "Hasta 60 caracteres; catálogo si aplica.",
            "El ID de riesgo no puede tener más de 60 caracteres.",
            "validators.validate_risk_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "Criticidad (riesgo catálogo)",
            "Obligatoria cuando se selecciona un riesgo del catálogo.",
            "Debe seleccionar una criticidad válida.",
            "ui/frames/risk.py",
            CLASSIFY_VALIDATION,
        ),
        (
            "Criticidad (riesgo catálogo)",
            "Obligatoria cuando se selecciona un riesgo del catálogo.",
            "Debe seleccionar una criticidad válida.",
            "ui/frames/risk.py",
        ),
        (
            "ID de cliente",
            "Valida longitud según tipo de documento.",
            "Mensaje específico según tipo de documento.",
            "validators.validate_client_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "ID de producto",
            "Valida longitud según tipo de producto (tarjeta/crédito/etc.).",
            "El ID del producto no cumple el formato requerido.",
            "validators.validate_product_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "ID de colaborador",
            "Formato letra + 5 dígitos.",
            "El ID del colaborador debe iniciar con una letra seguida de 5 dígitos.",
            "validators.validate_team_member_id",
            CLASSIFY_VALIDATION,
        ),
        (
            "N° de reclamos",
            "Debe ser numérico cuando se usa en el encabezado extendido.",
            "El número de reclamos debe ser numérico.",
            "app._validate_reclamos_count",
            CLASSIFY_VALIDATION,
        ),
        (
            "Analítica contable (encabezado)",
            "Código de 10 dígitos válido del catálogo.",
            "El código seleccionado no pertenece al catálogo de analíticas contables.",
            "app._validate_header_analitica",
            CLASSIFY_VALIDATION,
        ),
        (
            "Código de agencia",
            "6 dígitos numéricos.",
            "El código de agencia debe tener exactamente 6 dígitos.",
            "validators.validate_agency_code",
            CLASSIFY_VALIDATION,
        ),
        (
            "Reclamos obligatorios",
            "Si pérdida/falla/contingencia > 0, se exige reclamo completo.",
            "Debe ingresar al menos un reclamo completo porque hay montos de pérdida/falla/contingencia.",
            "ui/frames/products.py",
            CLASSIFY_VALIDATION,
        ),
        (
            "Agencia obligatoria",
            "Si división es DCA/Canales y área contiene 'area comercial', agencia es obligatoria.",
            "Debe ingresar nombre y código de agencia.",
            "ui/frames/team.py",
            CLASSIFY_VALIDATION,
        ),
        (
            "Involucramiento (ID/monto)",
            "Si hay ID se exige monto y viceversa.",
            "Debe completar ID y monto de involucramiento.",
            "ui/frames/products.py",
            CLASSIFY_VALIDATION,
        ),
        (
            "Involucramiento (ID/monto)",
            "Si hay ID se exige monto y viceversa.",
            "Debe completar ID y monto de involucramiento.",
            "ui/frames/products.py",
        ),
        (
            "Planes de acción",
            "IDs separados por ';' y sin duplicados entre riesgos.",
            "Plan de acción <id> duplicado entre riesgos.",
            "app.validate_data",
            CLASSIFY_VALIDATION,
        ),
        (
            "Llave técnica",
            "No permite duplicados con combinación caso+producto+cliente+colaborador+fecha+reclamo.",
            "Ya existe un registro con la misma llave técnica.",
            "utils/technical_key.py",
            CLASSIFY_VALIDATION,
        ),
    ]
    return [headers, *rows]


def _catalog_rows() -> list[Sequence[str | None]]:
    headers = (
        "Archivo de Catálogo/Detalle",
        "Columna en CSV",
        "Campo en la UI (Pestaña - Campo)",
        "Autocompletado",
        "Notas",
        "Clasificación",
    )
    rows = [
        ("client_details.csv", "id_cliente", "Clientes - ID del cliente", "Clave de búsqueda", "", CLASSIFY_LAKE),
        ("client_details.csv", "nombres", "Clientes - Nombres", "Sí", "", CLASSIFY_LAKE),
        ("client_details.csv", "apellidos", "Clientes - Apellidos", "Sí", "", CLASSIFY_LAKE),
        ("client_details.csv", "tipo_id", "Clientes - Tipo de ID", "Sí", "", CLASSIFY_LAKE),
        ("client_details.csv", "flag", "Clientes - Flag", "Sí", "", CLASSIFY_LAKE),
        ("client_details.csv", "telefonos", "Clientes - Teléfonos", "Sí", "", CLASSIFY_LAKE),
        ("client_details.csv", "correos", "Clientes - Correos", "Sí", "", CLASSIFY_LAKE),
        ("client_details.csv", "direcciones", "Clientes - Direcciones", "Sí", "", CLASSIFY_LAKE),
        ("client_details.csv", "accionado", "Clientes - Accionado", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "id_colaborador", "Colaboradores - ID del colaborador", "Clave de búsqueda", "", CLASSIFY_LAKE),
        ("team_details.csv", "nombres", "Colaboradores - Nombres", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "apellidos", "Colaboradores - Apellidos", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "flag", "Colaboradores - Flag", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "division", "Colaboradores - División", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "area", "Colaboradores - Área", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "servicio", "Colaboradores - Servicio", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "puesto", "Colaboradores - Puesto", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "fecha_carta_inmediatez", "Colaboradores - Fecha carta inmediatez", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "fecha_carta_renuncia", "Colaboradores - Fecha carta renuncia", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "nombre_agencia", "Colaboradores - Nombre agencia", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "codigo_agencia", "Colaboradores - Código agencia", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "tipo_falta", "Colaboradores - Tipo de falta", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "tipo_sancion", "Colaboradores - Tipo de sanción", "Sí", "", CLASSIFY_LAKE),
        ("team_details.csv", "fecha_actualizacion", "No visible en UI", "No", "Campo técnico del catálogo.", CLASSIFY_LAKE),
        ("risk_details.csv", "id_riesgo", "Riesgos - ID de riesgo", "Clave de búsqueda", "", CLASSIFY_LAKE),
        ("risk_details.csv", "lider", "Riesgos - Líder", "Sí", "", CLASSIFY_LAKE),
        ("risk_details.csv", "descripcion", "Riesgos - Descripción", "Sí", "", CLASSIFY_LAKE),
        ("risk_details.csv", "criticidad", "Riesgos - Criticidad", "Sí", "", CLASSIFY_LAKE),
        ("risk_details.csv", "exposicion_residual", "Riesgos - Exposición residual", "Sí", "", CLASSIFY_LAKE),
        ("risk_details.csv", "planes_accion", "Riesgos - Planes de acción", "Sí", "", CLASSIFY_LAKE),
        ("risk_details.csv", "pda", "No visible en UI", "No", "Dato auxiliar del catálogo.", CLASSIFY_LAKE),
        ("norm_details.csv", "id_norma", "Normas - ID de norma", "Clave de búsqueda", "", CLASSIFY_LAKE),
        ("norm_details.csv", "fecha_vigencia", "Normas - Fecha de vigencia", "Sí", "", CLASSIFY_LAKE),
        ("norm_details.csv", "descripcion", "Normas - Descripción", "Sí", "", CLASSIFY_LAKE),
        ("norm_details.csv", "acapite_inciso", "Normas - Acápite/Inciso", "Sí", "", CLASSIFY_LAKE),
        ("norm_details.csv", "detalle_norma", "Normas - Detalle de norma", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "id_producto", "Productos - ID del producto", "Clave de búsqueda", "Catálogo opcional.", CLASSIFY_LAKE),
        ("product_details.csv", "id_cliente", "Productos - Cliente", "Sí", "Autopobla titular.", CLASSIFY_LAKE),
        ("product_details.csv", "tipo_producto", "Productos - Tipo de producto", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "categoria1", "Productos - Categoría 1", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "categoria2", "Productos - Categoría 2", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "modalidad", "Productos - Modalidad", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "canal", "Productos - Canal", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "proceso", "Productos - Proceso", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "fecha_ocurrencia", "Productos - Fecha de ocurrencia", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "fecha_descubrimiento", "Productos - Fecha de descubrimiento", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "monto_investigado", "Productos - Monto investigado", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "tipo_moneda", "Productos - Moneda", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "monto_perdida_fraude", "Productos - Monto pérdida fraude", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "monto_falla_procesos", "Productos - Monto falla procesos", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "monto_contingencia", "Productos - Monto contingencia", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "monto_recuperado", "Productos - Monto recuperado", "Sí", "", CLASSIFY_LAKE),
        ("product_details.csv", "monto_pago_deuda", "Productos - Monto pago deuda", "Sí", "", CLASSIFY_LAKE),
        ("process_details.csv", "id_proceso", "Caso/Productos - ID Proceso", "Clave de búsqueda", "", CLASSIFY_LAKE),
        ("process_details.csv", "proceso", "Caso/Productos - Proceso impactado", "Sí", "", CLASSIFY_LAKE),
        ("process_details.csv", "canal", "Caso/Productos - Canal", "Sí", "", CLASSIFY_LAKE),
        ("process_details.csv", "descripcion", "Selector de procesos", "Sí", "Se muestra en selector.", CLASSIFY_LAKE),
        ("claim_details.csv", "id_reclamo", "Productos - ID reclamo", "Clave de búsqueda", "", CLASSIFY_LAKE),
        ("claim_details.csv", "nombre_analitica", "Productos - Analítica nombre", "Sí", "", CLASSIFY_LAKE),
        ("claim_details.csv", "codigo_analitica", "Productos - Código analítica", "Sí", "", CLASSIFY_LAKE),
        (
            "settings.TAXONOMIA",
            "-",
            "Caso/Productos - Taxonomía",
            "Catálogo interno",
            "Jerarquía categoría1 > categoría2 > modalidad.",
            CLASSIFY_LIST,
        ),
        ("settings.CANAL_LIST", "-", "Caso/Productos - Canal", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.PROCESO_LIST", "-", "Caso/Productos - Proceso", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.TIPO_PRODUCTO_LIST", "-", "Productos - Tipo de producto", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.TIPO_MONEDA_LIST", "-", "Productos - Moneda", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.TIPO_ID_LIST", "-", "Clientes - Tipo de ID", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.FLAG_CLIENTE_LIST", "-", "Clientes - Flag", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.ACCIONADO_OPTIONS", "-", "Clientes - Accionado", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.FLAG_COLABORADOR_LIST", "-", "Colaboradores - Flag", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.TIPO_FALTA_LIST", "-", "Colaboradores - Tipo de falta", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.TIPO_SANCION_LIST", "-", "Colaboradores - Tipo de sanción", "Catálogo interno", "", CLASSIFY_LIST),
        ("settings.CRITICIDAD_LIST", "-", "Riesgos - Criticidad", "Catálogo interno", "", CLASSIFY_LIST),
        ("TEAM_HIERARCHY_CATALOG", "-", "Colaboradores - División/Área/Servicio/Puesto", "Catálogo interno", "", CLASSIFY_LIST),
        ("AGENCY_CATALOG", "-", "Colaboradores - Agencia", "Catálogo interno", "", CLASSIFY_LIST),
        ("models.analitica_catalog", "-", "Productos/Análisis - Analítica contable", "Catálogo interno", "", CLASSIFY_LIST),
    ]
    return [headers, *rows]


def build_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = {
        "Caso y Participantes": _case_and_participants_rows(),
        "Riesgos": _risk_rows(),
        "Normas": _norm_rows(),
        "Análisis y Narrativas": _analysis_rows(),
        "Acciones": _actions_rows(),
        "Resumen": _summary_rows(),
        "Reportes y CSV": _export_structure_rows(),
        "Mapeo Exportes": _mapping_export_rows(),
        "Eventos_CSV": _eventos_rows(),
        "Logs": _logs_rows(),
        "Carta_inmediatez": _carta_rows(),
        "Informe_Gerencia": _gerencia_rows(),
        "Alerta_Temprana": _alerta_rows(),
        "Resumen_Ejecutivo": _resumen_ejecutivo_rows(),
        "Panel_Validacion": _validation_panel_rows(),
        "Mapeo Catalogos": _catalog_rows(),
    }

    for name, data in sheets.items():
        sheet = workbook.create_sheet(title=name)
        headers, *rows = data
        _apply_headers(sheet, headers)
        _append_rows(sheet, rows)
        for column_cells in sheet.columns:
            max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(80, max(16, max_len + 2))
    return workbook


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera el wireframe de Excel mejorado.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("docs/formulario_investigaciones_wireframe_mejorado.xlsx"),
        help="Ruta de salida del Excel mejorado.",
    )
    args = parser.parse_args()

    workbook = build_workbook()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Archivo generado en {args.output}")


if __name__ == "__main__":
    main()
