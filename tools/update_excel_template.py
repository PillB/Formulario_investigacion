"""Actualiza la plantilla normalizada de Excel con campos de normas y validaciones."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

from app import EXPORT_HEADERS, FraudCaseApp
from settings import EXTERNAL_DRIVE_DIR, EXTERNAL_LOGS_FILE

NORMAS_COLUMNS: Sequence[str] = tuple(EXPORT_HEADERS["detalles_norma.csv"])

NORMAS_DESCRIPTIONS: dict[str, tuple[str, str]] = {
    "id_norma": ("Clave primaria", "Identificador de la norma transgredida (formato XXXX.XXX.XX.XX)."),
    "id_caso": ("Clave foránea", "Identificador del caso; referencia CASOS.id_caso."),
    "descripcion": ("Atributo", "Descripción de la norma transgredida."),
    "fecha_vigencia": ("Atributo", "Fecha de vigencia de la norma (YYYY-MM-DD)."),
    "acapite_inciso": ("Atributo", "Referencia del acápite o inciso aplicable."),
    "detalle_norma": ("Atributo", "Amplía la explicación de la transgresión."),
}

NORMAS_VALIDATIONS: Sequence[tuple[str, str, str, str]] = (
    ("DETALLES_NORMA", "id_norma", "validate_norm_id", "Formato requerido: XXXX.XXX.XX.XX."),
    (
        "DETALLES_NORMA",
        "fecha_vigencia",
        "validate_date_text",
        "Formato YYYY-MM-DD; no debe ser futura.",
    ),
    (
        "DETALLES_NORMA",
        "acapite_inciso",
        "validate_required_text",
        "Campo requerido para registrar la norma.",
    ),
    (
        "DETALLES_NORMA",
        "detalle_norma",
        "validate_required_text",
        "Campo requerido; detalle narrativo de la transgresión.",
    ),
)

PANEL_VALIDATION_HEADERS: Sequence[str] = (
    "Regla",
    "Campo",
    "Pestaña",
    "Mensaje",
    "Lógica de validación",
)

PANEL_VALIDATION_ROWS: Sequence[tuple[str, str, str, str, str]] = (
    (
        "Disparo post-edición (FieldValidator)",
        "Campos con FieldValidator",
        "Todas",
        "Valida tras edición (focus out/selección); el modal aparece en focus out.",
        (
            "FieldValidator se arma con trace_add('write') y valida en <FocusOut>, "
            "<<ComboboxSelected>>, <<Paste>> o <<Cut>>. Usa debounce y evita validar "
            "hasta que el usuario edite el campo."
        ),
    ),
    (
        "Disparo post-edición (PostEditValidator)",
        "Campos del encabezado extendido",
        "Análisis y narrativas",
        "Valida tras edición (focus out/selección).",
        (
            "app._PostEditValidator se arma en KeyRelease y valida al perder foco o "
            "seleccionar un combobox; muestra messagebox si hay error."
        ),
    ),
    (
        "Formato de número de caso",
        "Número de caso",
        "Caso y participantes",
        "Debe ingresar el número de caso. / El número de caso debe seguir el formato AAAA-NNNN.",
        "validate_case_id en validators.py.",
    ),
    (
        "Formato de ID de proceso",
        "ID Proceso",
        "Caso y participantes",
        "Debe ingresar el ID de proceso. / El ID de proceso debe seguir el formato BPID-XXXXXX o BPID-RNF-XXXXXX.",
        "validate_process_id + normalización y autocompletado en app._validate_process_identifier.",
    ),
    (
        "Tipo de informe obligatorio",
        "Tipo de informe",
        "Caso y participantes",
        "Debe ingresar el tipo de informe.",
        "validate_required_text sobre tipo_informe_var.",
    ),
    (
        "Categoría nivel 1 obligatoria",
        "Categoría nivel 1",
        "Caso y participantes",
        "Debe ingresar la categoría nivel 1.",
        "validate_required_text sobre cat_caso1_var.",
    ),
    (
        "Categoría nivel 2 obligatoria",
        "Categoría nivel 2",
        "Caso y participantes",
        "Debe ingresar la categoría nivel 2.",
        "validate_required_text sobre cat_caso2_var.",
    ),
    (
        "Modalidad obligatoria",
        "Modalidad",
        "Caso y participantes",
        "Debe ingresar la modalidad del caso.",
        "validate_required_text sobre mod_caso_var.",
    ),
    (
        "Canal obligatorio",
        "Canal",
        "Caso y participantes",
        "Debe ingresar el canal del caso.",
        "validate_required_text sobre canal_caso_var.",
    ),
    (
        "Proceso impactado obligatorio",
        "Proceso impactado",
        "Caso y participantes",
        "Debe ingresar el proceso impactado.",
        "validate_required_text sobre proceso_caso_var.",
    ),
    (
        "Fecha de ocurrencia del caso",
        "Fecha de ocurrencia del caso",
        "Caso y participantes",
        (
            "La fecha de ocurrencia del caso debe tener el formato YYYY-MM-DD. "
            "No puede estar en el futuro y debe ser anterior a la fecha de descubrimiento."
        ),
        "validate_date_text con enforce_max_today y must_be_before.",
    ),
    (
        "Fecha de descubrimiento del caso",
        "Fecha de descubrimiento del caso",
        "Caso y participantes",
        (
            "La fecha de descubrimiento del caso debe tener el formato YYYY-MM-DD. "
            "No puede estar en el futuro y debe ser posterior a la fecha de ocurrencia."
        ),
        "validate_date_text con enforce_max_today y must_be_after.",
    ),
    (
        "Centro de costos",
        "Centro de costos",
        "Análisis y narrativas",
        "Cada centro de costos debe ser numérico y tener al menos 5 dígitos.",
        "app._validate_cost_centers con separadores ';'.",
    ),
    (
        "Fecha de reporte (opcional)",
        "Fecha de reporte",
        "Análisis y narrativas",
        "La fecha de reporte debe tener el formato YYYY-MM-DD y no puede estar en el futuro.",
        "validate_date_text (allow_blank=True, enforce_max_today=True).",
    ),
    (
        "Número de reclamos (opcional)",
        "N° de reclamos",
        "Análisis y narrativas",
        "El número de reclamos debe ser numérico.",
        "app._validate_reclamos_count.",
    ),
    (
        "Analítica contable",
        "Analítica contable",
        "Análisis y narrativas",
        (
            "El código de analítica debe tener 10 dígitos y comenzar con 43, 45, 46 o 56. "
            "El código seleccionado debe existir en el catálogo."
        ),
        "validate_codigo_analitica + find_analitica_by_code en app._validate_header_analitica.",
    ),
    (
        "Tipo de ID obligatorio",
        "Tipo de ID",
        "Caso y participantes",
        "Debe ingresar el tipo de ID del cliente.",
        "validate_required_text sobre tipo_id_var.",
    ),
    (
        "ID de cliente",
        "ID de cliente",
        "Caso y participantes",
        (
            "Debe ingresar el ID del cliente. "
            "DNI: 8 dígitos; RUC: 11 dígitos; Pasaporte/Carné: 9-12 alfanuméricos; "
            "otros tipos: mínimo 4 caracteres."
        ),
        "validate_client_id con patrones por tipo.",
    ),
    (
        "Nombres de cliente obligatorios",
        "Nombres",
        "Caso y participantes",
        "Debe ingresar los nombres del cliente.",
        "validate_required_text sobre nombres_var.",
    ),
    (
        "Apellidos de cliente obligatorios",
        "Apellidos",
        "Caso y participantes",
        "Debe ingresar los apellidos del cliente.",
        "validate_required_text sobre apellidos_var.",
    ),
    (
        "Flag de cliente",
        "Flag",
        "Caso y participantes",
        "Debe ingresar el flag del cliente. / El flag de cliente no está en el catálogo CM.",
        "validate_required_text + validación contra FLAG_CLIENTE_LIST.",
    ),
    (
        "Teléfonos del cliente",
        "Teléfonos",
        "Caso y participantes",
        "Debe ingresar al menos un teléfono del cliente. / Teléfono inválido.",
        "validate_required_text + validate_phone_list.",
    ),
    (
        "Correos del cliente",
        "Correos",
        "Caso y participantes",
        "Debe ingresar al menos un correo del cliente. / Correo inválido.",
        "validate_required_text + validate_email_list.",
    ),
    (
        "Accionado",
        "Accionado",
        "Caso y participantes",
        "Debe seleccionar al menos una opción en Accionado.",
        "validate_multi_selection.",
    ),
    (
        "ID de colaborador",
        "ID de Team member",
        "Caso y participantes",
        "Debe ingresar el ID del colaborador. / El ID del colaborador debe iniciar con una letra seguida de 5 dígitos.",
        "validate_team_member_id.",
    ),
    (
        "Nombres del colaborador obligatorios",
        "Nombres",
        "Caso y participantes",
        "Debe ingresar los nombres del colaborador.",
        "validate_required_text sobre nombres_var.",
    ),
    (
        "Apellidos del colaborador obligatorios",
        "Apellidos",
        "Caso y participantes",
        "Debe ingresar los apellidos del colaborador.",
        "validate_required_text sobre apellidos_var.",
    ),
    (
        "Flag del colaborador",
        "Flag",
        "Caso y participantes",
        "Debe seleccionar el flag del colaborador. / El flag no está en el catálogo CM.",
        "validate_required_text + validación contra FLAG_COLABORADOR_LIST.",
    ),
    (
        "División del colaborador",
        "División",
        "Caso y participantes",
        "Debe seleccionar la división del colaborador. / La división no está en el catálogo CM.",
        "TeamMemberFrame._validate_location_field('division').",
    ),
    (
        "Área del colaborador",
        "Área",
        "Caso y participantes",
        "Debe seleccionar el área del colaborador. / Selecciona la división antes de validar el área.",
        "TeamMemberFrame._validate_location_field('area').",
    ),
    (
        "Servicio del colaborador",
        "Servicio",
        "Caso y participantes",
        "Debe seleccionar el servicio del colaborador. / Completa división y área antes de validar el servicio.",
        "TeamMemberFrame._validate_location_field('servicio').",
    ),
    (
        "Puesto del colaborador",
        "Puesto",
        "Caso y participantes",
        "Debe seleccionar el puesto del colaborador. / Completa división, área y servicio antes de validar el puesto.",
        "TeamMemberFrame._validate_location_field('puesto').",
    ),
    (
        "Nombre de agencia (condicional)",
        "Nombre de agencia",
        "Caso y participantes",
        "Debe ingresar el nombre de la agencia.",
        (
            "TeamMemberFrame._validate_agency_fields('nombre'): requerido si división contiene "
            "'DCA' o 'Canales de atención' y área contiene 'area comercial'."
        ),
    ),
    (
        "Código de agencia (condicional)",
        "Código de agencia",
        "Caso y participantes",
        "El código de agencia debe tener exactamente 6 dígitos.",
        (
            "validate_agency_code en TeamMemberFrame._validate_agency_fields('codigo') "
            "con obligatoriedad condicional."
        ),
    ),
    (
        "Consistencia catálogo de agencias",
        "Nombre/Código de agencia",
        "Caso y participantes",
        "El nombre/código de agencia no coincide con el catálogo CM.",
        "TeamMemberFrame._validate_agency_fields valida coincidencia con team_details.",
    ),
    (
        "Fechas de cartas",
        "Fecha de carta de inmediatez / renuncia",
        "Caso y participantes",
        "La fecha debe tener el formato YYYY-MM-DD y no puede estar en el futuro.",
        "validate_date_text (allow_blank=True, enforce_max_today=True).",
    ),
    (
        "Tipo de falta",
        "Tipo de falta",
        "Caso y participantes",
        "Debe seleccionar el tipo de falta del colaborador. / El valor no está en el catálogo CM.",
        "TeamMemberFrame._validate_catalog_selection contra TIPO_FALTA_LIST.",
    ),
    (
        "Tipo de sanción",
        "Tipo de sanción",
        "Caso y participantes",
        "Debe seleccionar el tipo de sanción del colaborador. / El valor no está en el catálogo CM.",
        "TeamMemberFrame._validate_catalog_selection contra TIPO_SANCION_LIST.",
    ),
    (
        "ID de producto",
        "ID de producto",
        "Caso y participantes",
        (
            "Debe ingresar el ID del producto. "
            "Tarjetas: 16 alfanuméricos; Créditos: 13/14/16/20 alfanuméricos; "
            "Cuentas/depositos: 10-18 dígitos o 20 alfanuméricos con letra; "
            "otros: 4-30 alfanuméricos."
        ),
        "validate_product_id según tipo_producto.",
    ),
    (
        "Cliente del producto",
        "Cliente asociado",
        "Caso y participantes",
        "Debe ingresar el cliente del producto.",
        "validate_required_text sobre client_var.",
    ),
    (
        "Categoría 1 del producto",
        "Categoría 1",
        "Caso y participantes",
        "Debe ingresar la categoría 1.",
        "validate_required_text sobre cat1_var.",
    ),
    (
        "Categoría 2 del producto",
        "Categoría 2",
        "Caso y participantes",
        "Debe ingresar la categoría 2.",
        "validate_required_text sobre cat2_var.",
    ),
    (
        "Modalidad del producto",
        "Modalidad",
        "Caso y participantes",
        "Debe ingresar la modalidad.",
        "validate_required_text sobre mod_var.",
    ),
    (
        "Tipo de producto",
        "Tipo de producto",
        "Caso y participantes",
        "Debe ingresar el tipo de producto.",
        "validate_required_text sobre tipo_prod_var.",
    ),
    (
        "Catálogo de canal",
        "Canal del producto",
        "Caso y participantes",
        "Debe seleccionar el canal del producto. / El valor no está en el catálogo CM de canales.",
        "ProductFrame._validate_catalog_selection con CANAL_LIST.",
    ),
    (
        "Catálogo de proceso",
        "Proceso del producto",
        "Caso y participantes",
        "Debe seleccionar el proceso del producto. / El valor no está en el catálogo CM de procesos.",
        "ProductFrame._validate_catalog_selection con PROCESO_LIST.",
    ),
    (
        "Catálogo de moneda",
        "Moneda del producto",
        "Caso y participantes",
        "Debe seleccionar la moneda del producto. / El valor no está en el catálogo CM de tipos de moneda.",
        "ProductFrame._validate_catalog_selection con TIPO_MONEDA_LIST.",
    ),
    (
        "Fecha de ocurrencia del producto",
        "Fecha de ocurrencia",
        "Caso y participantes",
        (
            "La fecha de ocurrencia debe tener el formato YYYY-MM-DD, no estar en el futuro "
            "y ser anterior a la fecha de descubrimiento."
        ),
        "validate_date_text + validate_product_dates.",
    ),
    (
        "Fecha de descubrimiento del producto",
        "Fecha de descubrimiento",
        "Caso y participantes",
        (
            "La fecha de descubrimiento debe tener el formato YYYY-MM-DD, no estar en el futuro "
            "y ser posterior a la fecha de ocurrencia."
        ),
        "validate_date_text + validate_product_dates.",
    ),
    (
        "Monto investigado",
        "Monto investigado",
        "Caso y participantes",
        (
            "Debe ingresar el monto investigado. Debe ser numérico, >= 0, máximo 12 dígitos enteros "
            "y hasta 2 decimales."
        ),
        "validate_money_bounds (allow_blank=False).",
    ),
    (
        "Montos del producto",
        "Pérdida/Falla/Contingencia/Recuperado/Pago deuda",
        "Caso y participantes",
        "Debe ser numérico, >= 0, máximo 12 dígitos enteros y hasta 2 decimales.",
        "validate_money_bounds (allow_blank=True) en ProductFrame._validate_amount_input.",
    ),
    (
        "Consistencia de montos",
        "Montos del producto",
        "Caso y participantes",
        (
            "La suma de pérdida, falla, contingencia y recuperación debe igualar el monto investigado. "
            "El recuperado y el pago de deuda no pueden superar el investigado."
        ),
        "ProductFrame._validate_montos_consistentes.",
    ),
    (
        "Regla contingencia crédito/tarjeta",
        "Monto contingencia",
        "Caso y participantes",
        "El monto de contingencia debe ser igual al monto investigado para créditos o tarjetas.",
        "ProductFrame._validate_montos_consistentes con tipo_producto.",
    ),
    (
        "Reclamo requerido por montos",
        "Reclamos asociados",
        "Caso y participantes",
        (
            "Debe ingresar al menos un reclamo completo si hay montos de pérdida, falla o contingencia."
        ),
        "ProductFrame._claim_fields_required y _show_claim_requirement_error.",
    ),
    (
        "ID de reclamo",
        "ID de reclamo",
        "Caso y participantes",
        "El ID de reclamo debe tener el formato CXXXXXXXX.",
        "validate_reclamo_id en ClaimRow.",
    ),
    (
        "Nombre de analítica",
        "Nombre de analítica",
        "Caso y participantes",
        "Debe ingresar el nombre de la analítica / No existe en el catálogo.",
        "validate_required_text + _validate_analitica_catalog (find_analitica_by_name).",
    ),
    (
        "Código de analítica",
        "Código de analítica",
        "Caso y participantes",
        "El código de analítica debe tener 10 dígitos y comenzar con 43, 45, 46 o 56.",
        "validate_codigo_analitica + _validate_analitica_catalog (find_analitica_by_code).",
    ),
    (
        "Consistencia analítica",
        "Nombre/Código de analítica",
        "Caso y participantes",
        "El nombre y el código de analítica no corresponden al mismo elemento del catálogo.",
        "ClaimRow._validate_analitica_catalog.",
    ),
    (
        "Involucramiento de colaboradores",
        "Colaborador involucrado",
        "Caso y participantes",
        "Debe seleccionar un colaborador válido y registrar el monto asignado.",
        "AssignmentRow._validate_selection con IDs existentes.",
    ),
    (
        "Involucramiento de clientes",
        "Cliente involucrado",
        "Caso y participantes",
        "Debe seleccionar un cliente válido y registrar el monto asignado.",
        "AssignmentRow._validate_selection con IDs existentes.",
    ),
    (
        "Monto asignado",
        "Monto asignado",
        "Caso y participantes",
        "Debe ser numérico, >= 0, máximo 12 dígitos enteros y hasta 2 decimales.",
        "AssignmentRow._validate_assignment_amount (validate_money_bounds).",
    ),
    (
        "ID de riesgo",
        "ID de riesgo",
        "Riesgos",
        (
            "Nuevo riesgo: máximo 60 caracteres imprimibles. "
            "Catálogo: formato RSK-########."
        ),
        "validate_risk_id o validate_catalog_risk_id según modo.",
    ),
    (
        "Líder del riesgo",
        "Líder",
        "Riesgos",
        "Debe ingresar el líder del riesgo (modo catálogo).",
        "RiskFrame._validate_when_catalog con validate_required_text.",
    ),
    (
        "Exposición residual",
        "Exposición residual",
        "Riesgos",
        "Debe ser numérica, >= 0, máximo 12 dígitos enteros y hasta 2 decimales.",
        "validate_money_bounds (allow_blank=True).",
    ),
    (
        "Criticidad",
        "Criticidad",
        "Riesgos",
        "Debe seleccionar la criticidad del riesgo. / La criticidad no está en el catálogo CM.",
        "RiskFrame._validate_criticidad.",
    ),
    (
        "Descripción del riesgo",
        "Descripción",
        "Riesgos",
        "Debe ingresar la descripción del riesgo.",
        "validate_required_text sobre descripcion_var.",
    ),
    (
        "Planes de acción",
        "Planes de acción",
        "Riesgos",
        "Debe ingresar los planes de acción (modo catálogo).",
        "RiskFrame._validate_when_catalog con validate_required_text.",
    ),
    (
        "ID de norma",
        "ID de norma",
        "Normas",
        "El ID de norma debe seguir el formato XXXX.XXX.XX.XX.",
        "validate_norm_id.",
    ),
    (
        "Fecha de vigencia de norma",
        "Fecha de vigencia",
        "Normas",
        "Debe tener el formato YYYY-MM-DD y no puede estar en el futuro.",
        "validate_date_text (allow_blank=False, enforce_max_today=True).",
    ),
    (
        "Descripción de norma",
        "Descripción",
        "Normas",
        "Debe ingresar la descripción de la norma.",
        "validate_required_text sobre descripcion_var.",
    ),
    (
        "Acápite/Inciso de norma",
        "Acápite/Inciso",
        "Normas",
        "Debe ingresar el acápite o inciso de la norma.",
        "validate_required_text sobre acapite_var.",
    ),
    (
        "Detalle de norma",
        "Detalle de norma",
        "Normas",
        "Debe ingresar el detalle de la norma.",
        "validate_required_text sobre detalle_var / texto enriquecido.",
    ),
    (
        "Validación de clave técnica (duplicados)",
        "Clave técnica",
        "Panel de validación",
        (
            "Bloquea si falta fecha de ocurrencia o asociaciones, si la fecha es inválida "
            "o si se detecta duplicado en la combinación caso-producto-cliente-colaborador-fecha-reclamo."
        ),
        "app._check_duplicate_technical_keys_realtime + build_technical_key.",
    ),
)

EXTERNAL_DRIVE_HEADERS: Sequence[str] = ("Ubicación", "Archivo/Patrón", "Reglas de copia")
EXTERNAL_DRIVE_ROWS: Sequence[tuple[str, str, str]] = (
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}",
        Path(EXTERNAL_LOGS_FILE).name,
        (
            "Se escribe al vaciar la cola de logs si la unidad externa está disponible. "
            "app._flush_log_queue_to_disk usa _resolve_external_log_target para anexar."
        ),
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        (
            "<prefijo>_casos.csv, _clientes.csv, _colaboradores.csv, _productos.csv, "
            "_producto_reclamo.csv, _involucramiento.csv, _detalles_riesgo.csv, "
            "_detalles_norma.csv, _llave_tecnica.csv, _eventos.csv, _analisis.csv, _logs.csv"
        ),
        (
            "Se copian al usar Guardar y enviar. _perform_save_exports construye los CSV y "
            "_mirror_exports_to_external_drive los duplica (se omite autosave.json)."
        ),
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        "<prefijo>_version.json",
        "Se copia junto con los CSV al usar Guardar y enviar.",
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        "<nombre_informe>.md",
        (
            "Se duplica al exportar Guardar y enviar (save_md) y al usar Generar informe Markdown."
        ),
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        "<nombre_informe>.docx",
        (
            "Se duplica al exportar Guardar y enviar y al usar Generar informe Word, "
            "solo si python-docx está disponible."
        ),
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        "<nombre_informe>.pptx",
        (
            "Se duplica al usar Generar alerta temprana (PPTX), "
            "solo si python-pptx está disponible."
        ),
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        "architecture.mmd",
        "Se copia junto con las exportaciones cuando se actualiza el diagrama.",
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        "h_<tabla>.csv",
        (
            "Se copia junto con las exportaciones; si falla la copia o no hay unidad externa, "
            "se registra en pending_consolidation.txt para reintento en el siguiente inicio."
        ),
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}/<id_caso>",
        "<id_caso>_temp_<YYYYMMDD_HHMMSS>.json",
        (
            "Se escribe al guardar versiones temporales (save_temp_version) cuando hay unidad externa."
        ),
    ),
    (
        f"{Path(EXTERNAL_DRIVE_DIR).as_posix()}",
        "h_cartas_inmediatez.csv",
        "Se actualiza al generar cartas de inmediatez si la unidad externa está disponible.",
    ),
)

SHEET_HEADERS: dict[str, Sequence[str]] = {
    "CASOS": EXPORT_HEADERS["casos.csv"],
    "CLIENTES": EXPORT_HEADERS["clientes.csv"],
    "COLABORADORES": EXPORT_HEADERS["colaboradores.csv"],
    "PRODUCTOS": EXPORT_HEADERS["productos.csv"],
    "PRODUCTO_RECLAMO": EXPORT_HEADERS["producto_reclamo.csv"],
    "INVOLUCRAMIENTO": EXPORT_HEADERS["involucramiento.csv"],
    "DETALLES_RIESGO": EXPORT_HEADERS["detalles_riesgo.csv"],
    "DETALLES_NORMA": EXPORT_HEADERS["detalles_norma.csv"],
    "ANALISIS": EXPORT_HEADERS["analisis.csv"],
}


def _set_header_row(sheet, columns: Sequence[str]) -> None:
    for idx, name in enumerate(columns, start=1):
        sheet.cell(row=1, column=idx, value=name)
    for idx in range(len(columns) + 1, sheet.max_column + 1):
        sheet.cell(row=1, column=idx, value=None)


def _rewrite_rows(sheet, rows: Iterable[Sequence[str]]) -> None:
    max_row = sheet.max_row
    if max_row and max_row > 1:
        sheet.delete_rows(2, max_row - 1)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)


def _update_description_sheet(sheet) -> None:
    existing_rows = [
        row
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] != "DETALLES_NORMA"
    ]
    normas_rows = [
        ("DETALLES_NORMA", col, *NORMAS_DESCRIPTIONS[col]) for col in NORMAS_COLUMNS
    ]
    _rewrite_rows(sheet, [*existing_rows, *normas_rows])


def _update_validation_sheet(workbook, rows: Sequence[tuple[str, str, str, str]]) -> None:
    if "VALIDACIONES" in workbook.sheetnames:
        sheet = workbook["VALIDACIONES"]
    else:
        sheet = workbook.create_sheet(title="VALIDACIONES")
    sheet.cell(row=1, column=1, value="Hoja")
    sheet.cell(row=1, column=2, value="Columna")
    sheet.cell(row=1, column=3, value="Validador")
    sheet.cell(row=1, column=4, value="Regla")

    existing_rows = [
        row
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and row[0] != "DETALLES_NORMA"
    ]
    _rewrite_rows(sheet, [*existing_rows, *rows])


def _update_validation_panel_sheet(workbook) -> None:
    if "Panel de validación" in workbook.sheetnames:
        sheet = workbook["Panel de validación"]
    else:
        sheet = workbook.create_sheet(title="Panel de validación")
    _set_header_row(sheet, PANEL_VALIDATION_HEADERS)
    _rewrite_rows(sheet, PANEL_VALIDATION_ROWS)


def _update_external_drive_sheet(workbook) -> None:
    if "Logs/External Drive" in workbook.sheetnames:
        sheet = workbook["Logs/External Drive"]
    else:
        sheet = workbook.create_sheet(title="Logs/External Drive")
    _set_header_row(sheet, EXTERNAL_DRIVE_HEADERS)
    _rewrite_rows(sheet, EXTERNAL_DRIVE_ROWS)


def _update_summary_sheet(workbook) -> None:
    summary_config = FraudCaseApp.build_summary_table_config()
    if "RESUMEN" in workbook.sheetnames:
        sheet = workbook["RESUMEN"]
    else:
        sheet = workbook.create_sheet(title="RESUMEN")
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    row_idx = 1
    for _key, title, columns in summary_config:
        sheet.cell(row=row_idx, column=1, value=title)
        row_idx += 1
        for col_idx, (_field, label) in enumerate(columns, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=label)
        row_idx += 2


def update_template(path: Path) -> None:
    workbook = load_workbook(path)
    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        _set_header_row(sheet, headers)

    description_sheet = workbook["DESCRIPCION_COLUMNAS"]
    _update_description_sheet(description_sheet)
    _update_validation_sheet(workbook, NORMAS_VALIDATIONS)
    _update_validation_panel_sheet(workbook)
    _update_external_drive_sheet(workbook)
    _update_summary_sheet(workbook)

    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Actualiza la plantilla normalizada con campos de normas y validaciones."
    )
    parser.add_argument(
        "--path",
        type=Path,
        default=Path("plantilla_normalizada_convalidaciones.xlsx"),
        help="Ruta del archivo Excel a actualizar.",
    )
    args = parser.parse_args()
    update_template(args.path)


if __name__ == "__main__":
    main()
