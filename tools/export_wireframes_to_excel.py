"""Exporta un wireframe aproximado de la interfaz Tkinter a Excel.

El script instancia la aplicación real, recorre los widgets colocados en las
seis pestañas principales y los proyecta a una cuadrícula de Excel usando
``openpyxl``. Cada celda queda etiquetada con el tipo de control (texto,
número, selección única, multiselección o autocompletar), su estado (por
ejemplo, readonly) y la sección a la que pertenece. Se aplican colores y una
leyenda para distinguir rápidamente campos, botones y badges de validación.

Ejemplo de uso::

    python tools/export_wireframes_to_excel.py \
        --output wireframes/Formulario_UI_wireframe.xlsx

El recorrido evita modificar datos persistentes utilizando el modo de arranque
de pruebas y seleccionando la opción de "nuevo caso" para construir la UI sin
diálogos.
"""

from __future__ import annotations

import argparse
import math
import os
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, Mapping, Sequence

import tkinter as tk
from tkinter import TclError, scrolledtext, ttk

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

from app import (
    COMENTARIO_AMPLIO_MAX_CHARS,
    COMENTARIO_BREVE_MAX_CHARS,
    FraudCaseApp,
)
from report_builder import build_event_rows, build_llave_tecnica_rows, build_report_filename
from validators import LOG_FIELDNAMES, normalize_log_row
from validation_badge.validation_badge import (
    NEUTRAL_STYLE,
    SUCCESS_STYLE,
    WARNING_STYLE,
    ValidationBadge,
)


CELL_WIDTH_PX = 48
CELL_HEIGHT_PX = 26

STYLE_MAP = {
    "text": PatternFill("solid", fgColor="DDEBF7"),
    "number": PatternFill("solid", fgColor="FFE699"),
    "single-select": PatternFill("solid", fgColor="E2EFDA"),
    "multi-select": PatternFill("solid", fgColor="DDD9C3"),
    "autocomplete": PatternFill("solid", fgColor="C6E0B4"),
    "action": PatternFill("solid", fgColor="D9D9D9"),
    "badge": PatternFill("solid", fgColor="F8CBAD"),
    "label": PatternFill("solid", fgColor="F2F2F2"),
}

REPO_ROOT = Path(__file__).resolve().parents[1]

ANALYSIS_HEADERS: Sequence[str] = (
    "id_caso",
    "antecedentes",
    "modus_operandi",
    "hallazgos",
    "descargos",
    "conclusiones",
    "recomendaciones",
    "comentario_breve",
    "comentario_amplio",
)

EXPORT_HEADER_MAP: Sequence[tuple[str, Sequence[str] | None]] = (
    (
        "casos.csv",
        (
            "id_caso",
            "tipo_informe",
            "categoria1",
            "categoria2",
            "modalidad",
            "canal",
            "proceso",
            "fecha_de_ocurrencia",
            "fecha_de_descubrimiento",
            "centro_costos",
            "matricula_investigador",
            "investigador_nombre",
            "investigador_cargo",
        ),
    ),
    ("llave_tecnica.csv", None),
    ("eventos.csv", None),
    (
        "clientes.csv",
        (
            "id_cliente",
            "id_caso",
            "nombres",
            "apellidos",
            "tipo_id",
            "flag",
            "telefonos",
            "correos",
            "direcciones",
            "accionado",
        ),
    ),
    (
        "colaboradores.csv",
        (
            "id_colaborador",
            "id_caso",
            "flag",
            "nombres",
            "apellidos",
            "division",
            "area",
            "servicio",
            "puesto",
            "fecha_carta_inmediatez",
            "fecha_carta_renuncia",
            "nombre_agencia",
            "codigo_agencia",
            "tipo_falta",
            "tipo_sancion",
        ),
    ),
    (
        "productos.csv",
        (
            "id_producto",
            "id_caso",
            "id_cliente",
            "categoria1",
            "categoria2",
            "modalidad",
            "canal",
            "proceso",
            "fecha_ocurrencia",
            "fecha_descubrimiento",
            "monto_investigado",
            "tipo_moneda",
            "monto_perdida_fraude",
            "monto_falla_procesos",
            "monto_contingencia",
            "monto_recuperado",
            "monto_pago_deuda",
            "tipo_producto",
        ),
    ),
    (
        "producto_reclamo.csv",
        (
            "id_reclamo",
            "id_caso",
            "id_producto",
            "nombre_analitica",
            "codigo_analitica",
        ),
    ),
    (
        "involucramiento.csv",
        (
            "id_producto",
            "id_caso",
            "id_colaborador",
            "monto_asignado",
        ),
    ),
    (
        "detalles_riesgo.csv",
        (
            "id_riesgo",
            "id_caso",
            "lider",
            "descripcion",
            "criticidad",
            "exposicion_residual",
            "planes_accion",
        ),
    ),
    (
        "detalles_norma.csv",
        (
            "id_norma",
            "id_caso",
            "descripcion",
            "fecha_vigencia",
            "acapite_inciso",
            "detalle_norma",
        ),
    ),
    ("analisis.csv", ANALYSIS_HEADERS),
    ("logs.csv", LOG_FIELDNAMES),
)

MASSIVE_FILES: Sequence[tuple[str, str]] = (
    ("clientes_masivos.csv", "Clientes masivos"),
    ("colaboradores_masivos.csv", "Colaboradores masivos"),
    ("productos_masivos.csv", "Productos masivos"),
    ("datos_combinados_masivos.csv", "Archivo combinado masivo"),
    ("riesgos_masivos.csv", "Riesgos masivos"),
    ("normas_masivas.csv", "Normas masivas"),
    ("reclamos_masivos.csv", "Reclamos masivos"),
)

EMPTY_CASE_DATA: Mapping[str, object] = {
    "caso": {},
    "productos": [],
    "reclamos": [],
    "involucramientos": [],
    "clientes": [],
    "colaboradores": [],
}

COMMENTARY_VALIDATIONS: Mapping[str, Mapping[str, object]] = {
    "comentario breve": {
        "max_chars": COMENTARIO_BREVE_MAX_CHARS,
        "allow_newlines": False,
        "auto_redaccion": (
            "Auto-redactar: resumen automático sin PII y sin saltos de línea "
            "(max_new_tokens=72)."
        ),
    },
    "comentario amplio": {
        "max_chars": COMENTARIO_AMPLIO_MAX_CHARS,
        "allow_newlines": False,
        "auto_redaccion": (
            "Auto-redactar: resumen automático sin PII y sin saltos de línea "
            "(max_new_tokens=320)."
        ),
    },
}


@dataclass
class WidgetRecord:
    label: str
    widget_type: str
    control_kind: str
    state: str | None
    sticky: str | None
    tooltip: str | None
    section_path: tuple[str, ...]
    x: int
    y: int
    width: int
    height: int
    grid_row: int | None
    grid_column: int | None
    grid_rowspan: int | None
    grid_colspan: int | None


def _safe_grid_info(widget) -> dict:
    try:
        info = widget.grid_info()
        return info or {}
    except Exception:
        return {}


def _normalize_label(label: str) -> str:
    return label.strip().rstrip(":").lower()


def _widget_text(widget) -> str:
    getter = getattr(widget, "cget", None)
    if callable(getter):
        try:
            text = getter("text")
        except Exception:
            text = ""
        if text:
            return str(text)
    return ""


def _infer_label_for_widget(widget) -> str:
    parent = getattr(widget, "master", None)
    if parent is None:
        return ""
    target_info = _safe_grid_info(widget)
    if not target_info:
        return ""
    try:
        target_row = int(target_info.get("row", 0))
        target_col = int(target_info.get("column", 0))
    except Exception:
        return ""

    for sibling in parent.winfo_children():
        if sibling is widget:
            continue
        if not isinstance(sibling, (ttk.Label, tk.Label)):
            continue
        info = _safe_grid_info(sibling)
        try:
            row = int(info.get("row", 0))
            col = int(info.get("column", 0))
        except Exception:
            continue
        if col == target_col and row in {target_row, target_row - 1}:
            text = _widget_text(sibling)
            if text:
                return text
    return ""


def _control_kind(widget, label: str) -> str:
    lower_label = label.lower()
    if isinstance(widget, ValidationBadge):
        return "badge"
    if isinstance(widget, ttk.Combobox):
        state = _safe_cget(widget, "state")
        return "single-select" if state == "readonly" else "autocomplete"
    if isinstance(widget, (ttk.Treeview,)):
        return "multi-select"
    if isinstance(widget, (tk.Text, scrolledtext.ScrolledText)):
        return "multi-select"
    if isinstance(widget, (ttk.Checkbutton, ttk.Radiobutton)):
        return "single-select"
    if isinstance(widget, ttk.Button):
        return "action"
    if isinstance(widget, (ttk.Label, tk.Label)):
        style = (_safe_cget(widget, "style") or "").strip()
        if style in {WARNING_STYLE, SUCCESS_STYLE, NEUTRAL_STYLE}:
            return "badge"
        return "label"
    if isinstance(widget, (ttk.Entry, tk.Entry)):
        if any(token in lower_label for token in ("monto", "importe", "pago", "%")):
            return "number"
        if any(token in lower_label for token in ("código", "codigo", "id", "número", "numero")):
            return "text"
        return "text"
    return "text"


def _safe_cget(widget, option: str) -> str | None:
    getter = getattr(widget, "cget", None)
    if callable(getter):
        try:
            return getter(option)
        except Exception:
            return None
    return None


def _resolve_state(widget) -> str | None:
    for option in ("state", "values"):
        value = _safe_cget(widget, option)
        if value not in (None, "", ()):  # type: ignore[comparison-overlap]
            if option == "values" and isinstance(widget, ttk.Combobox):
                continue
            return str(value)
    return None


def _measure_widget(widget, origin_x: int, origin_y: int) -> tuple[int, int, int, int]:
    x = max(0, widget.winfo_rootx() - origin_x)
    y = max(0, widget.winfo_rooty() - origin_y)
    width = widget.winfo_width() or widget.winfo_reqwidth()
    height = widget.winfo_height() or widget.winfo_reqheight()
    return x, y, width, height


def _iter_widget_records(
    root_widget,
    origin_x: int,
    origin_y: int,
    section_stack: tuple[str, ...],
    tooltip_map: Mapping[object, str],
) -> Iterator[WidgetRecord]:
    for child in root_widget.winfo_children():
        try:
            mapped = child.winfo_ismapped()
        except Exception:
            mapped = True
        if not mapped:
            continue

        if isinstance(child, (ttk.LabelFrame, tk.LabelFrame)):
            text = _widget_text(child)
            next_stack = (*section_stack, text) if text else section_stack
            yield from _iter_widget_records(child, origin_x, origin_y, next_stack, tooltip_map)
            continue

        text = _widget_text(child) or _infer_label_for_widget(child)
        is_relevant = isinstance(
            child,
            (
                ttk.Entry,
                tk.Entry,
                ttk.Combobox,
                tk.Text,
                scrolledtext.ScrolledText,
                ttk.Treeview,
                ttk.Button,
                ttk.Label,
                tk.Label,
                ttk.Checkbutton,
                ttk.Radiobutton,
                ValidationBadge,
            ),
        )
        if not is_relevant:
            yield from _iter_widget_records(child, origin_x, origin_y, section_stack, tooltip_map)
            continue

        x, y, width, height = _measure_widget(child, origin_x, origin_y)
        if width <= 1 or height <= 1:
            continue

        grid_info = _safe_grid_info(child)
        record = WidgetRecord(
            label=text or child.__class__.__name__,
            widget_type=child.__class__.__name__,
            control_kind=_control_kind(child, text),
            state=_resolve_state(child),
            sticky=grid_info.get("sticky"),
            tooltip=tooltip_map.get(child),
            section_path=section_stack,
            x=x,
            y=y,
            width=width,
            height=height,
            grid_row=int(grid_info.get("row", 0)) if grid_info else None,
            grid_column=int(grid_info.get("column", 0)) if grid_info else None,
            grid_rowspan=int(grid_info.get("rowspan", 1)) if grid_info else None,
            grid_colspan=int(grid_info.get("columnspan", 1)) if grid_info else None,
        )
        yield record
        yield from _iter_widget_records(child, origin_x, origin_y, section_stack, tooltip_map)


def _normalize_to_cells(record: WidgetRecord, *, row_offset: int) -> tuple[int, int, int, int]:
    row = row_offset + math.floor(record.y / CELL_HEIGHT_PX) + 1
    column = math.floor(record.x / CELL_WIDTH_PX) + 1
    row_span = max(1, math.ceil(record.height / CELL_HEIGHT_PX))
    column_span = max(1, math.ceil(record.width / CELL_WIDTH_PX))
    return row, column, row_span, column_span


def _add_legend(sheet, start_row: int = 1) -> int:
    sheet.cell(row=start_row, column=1, value="Leyenda de controles").font = Font(bold=True)
    row = start_row + 1
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for label, fill in STYLE_MAP.items():
        cell = sheet.cell(row=row, column=1, value=label)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        row += 1
    sheet.cell(row=row, column=1, value="Las celdas incluyen estado y sticky en comentarios.")
    return row + 1


def _place_records(sheet, records: Iterable[WidgetRecord], *, row_offset: int) -> None:
    merged_ranges = []
    max_col = 1
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for record in records:
        row, col, row_span, col_span = _normalize_to_cells(record, row_offset=row_offset)
        max_col = max(max_col, col + col_span)
        cell = sheet.cell(row=row, column=col, value=record.label)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
        style = STYLE_MAP.get(record.control_kind)
        if style:
            cell.fill = style
        comment_lines = [f"Widget: {record.widget_type}", f"Tipo de control: {record.control_kind}"]
        if record.state:
            comment_lines.append(f"Estado: {record.state}")
        if record.section_path:
            comment_lines.append(f"Sección: {' > '.join(filter(None, record.section_path))}")
        comment_lines.append(
            f"Grid(row={record.grid_row}, col={record.grid_column}, rowspan={record.grid_rowspan}, colspan={record.grid_colspan}, sticky={record.sticky})"
        )
        if record.tooltip:
            comment_lines.append(f"Descripción: {record.tooltip}")
        normalized_label = _normalize_label(record.label)
        validation = COMMENTARY_VALIDATIONS.get(normalized_label)
        if validation and record.widget_type in {"ScrolledText", "Text"}:
            comment_lines.append(
                "Validaciones: max_chars="
                f"{validation['max_chars']}, allow_newlines={validation['allow_newlines']}"
            )
            auto_note = validation.get("auto_redaccion")
            if auto_note:
                comment_lines.append(f"Acción: {auto_note}")
        cell.comment = Comment("\n".join(comment_lines), "wireframe")
        if row_span > 1 or col_span > 1:
            target = f"{get_column_letter(col)}{row}:{get_column_letter(col + col_span - 1)}{row + row_span - 1}"
            conflict = any(_ranges_overlap(rng.coord, target) for rng in merged_ranges)
            if not conflict:
                sheet.merge_cells(target)
                merged_ranges.append(sheet.merged_cells.ranges[-1])
    for column in range(1, max_col + 2):
        sheet.column_dimensions[get_column_letter(column)].width = 18


def _ranges_overlap(existing: str, target: str) -> bool:
    min_col, min_row, max_col, max_row = range_boundaries(existing)
    t_min_col, t_min_row, t_max_col, t_max_row = range_boundaries(target)
    rows_overlap = not (max_row < t_min_row or t_max_row < min_row)
    cols_overlap = not (max_col < t_min_col or t_max_col < min_col)
    return rows_overlap and cols_overlap


def _placeholder_value(field: str) -> str:
    lower = field.lower()
    if "fecha" in lower:
        return "2024-05-01 (YYYY-MM-DD <= hoy)"
    if "monto" in lower or "importe" in lower or "exposicion" in lower:
        return "12345.67 (2 decimales, 12 dígitos)"
    if lower == "id_caso":
        return "2024-0001 (AAAA-NNNN)"
    if "id_reclamo" in lower:
        return "C12345678 (C+8 dígitos)"
    if "id_colaborador" in lower:
        return "T12345 (una letra + 5 dígitos)"
    if "codigo_analitica" in lower:
        return "4300000001 (10 dígitos, 43/45/46/56)"
    if "codigo_agencia" in lower:
        return "123456 (6 dígitos)"
    if "telefonos" in lower:
        return "999999999;988888888"
    if "correos" in lower:
        return "a@ejemplo.com;b@ejemplo.com"
    if "tipo_moneda" in lower:
        return "PEN|USD"
    if "tipologia" in lower or "tipo_informe" in lower:
        return "Texto controlado"
    return "Texto libre"


def _build_sample_row(headers: Sequence[str]) -> list[str]:
    return [_placeholder_value(field) for field in headers]


def _build_row_from_mapping(headers: Sequence[str], row: Mapping[str, str]) -> list[str]:
    return [row.get(field, "") for field in headers]


def _write_schema_block(
    sheet,
    start_row: int,
    title: str,
    headers: Sequence[str],
    sample_row: Sequence[str],
    note: str,
) -> int:
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True)
    if note:
        title_cell.comment = Comment(note, "wireframe")

    header_row = start_row + 1
    for idx, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=header_row, column=idx, value=header)
        header_cell.font = Font(bold=True)

    sample_row_index = header_row + 1
    for idx, value in enumerate(sample_row, start=1):
        sheet.cell(row=sample_row_index, column=idx, value=value)

    return sample_row_index + 2


def _resolve_dynamic_headers(filename: str) -> Sequence[str]:
    if filename == "llave_tecnica.csv":
        return build_llave_tecnica_rows(EMPTY_CASE_DATA)[1]
    if filename == "eventos.csv":
        return build_event_rows(EMPTY_CASE_DATA)[1]
    return []


def _load_massive_headers(base_dir: Path) -> list[tuple[str, list[str], Path]]:
    headers: list[tuple[str, list[str], Path]] = []
    for filename, label in MASSIVE_FILES:
        path = base_dir / filename
        header: list[str] = []
        try:
            with path.open(newline='', encoding="utf-8-sig") as handle:
                reader = csv.DictReader(line for line in handle if line.strip())
                header = list(reader.fieldnames or [])
        except FileNotFoundError:
            header = []
        headers.append((label, header, path))
    return headers


def _add_docx_md_sheet(workbook) -> None:
    sheet = workbook.create_sheet(title="Reportes_DOCX_MD")
    docx_name = build_report_filename("Inicial", "2024-0001", "docx")
    md_name = build_report_filename("Inicial", "2024-0001", "md")
    sheet.cell(row=1, column=1, value="Ejemplos de nombres de archivo").font = Font(bold=True)
    sheet.cell(row=2, column=1, value=docx_name)
    sheet.cell(row=2, column=2, value=md_name)
    sheet.cell(row=3, column=1, value="Fuente: report_builder.build_report_filename")

    row = 5
    llave_header = build_llave_tecnica_rows(EMPTY_CASE_DATA)[1]
    row = _write_schema_block(
        sheet,
        row,
        "Llave técnica (build_llave_tecnica_rows)",
        llave_header,
        _build_sample_row(llave_header),
        "Campos usados en DOCX/MD y CSV; fuente: report_builder.build_llave_tecnica_rows",
    )

    event_header = build_event_rows(EMPTY_CASE_DATA)[1]
    row = _write_schema_block(
        sheet,
        row,
        "Eventos (build_event_rows)",
        event_header,
        _build_sample_row(event_header),
        "Orden real de exportación desde report_builder.build_event_rows",
    )

    row = _write_schema_block(
        sheet,
        row,
        "Narrativas (build_analysis_tab)",
        ANALYSIS_HEADERS,
        _build_sample_row(ANALYSIS_HEADERS),
        "Campos enriquecidos de análisis; origen: app.build_analysis_tab",
    )


def _add_csv_exports_sheet(workbook, base_dir: Path) -> None:
    sheet = workbook.create_sheet(title="Exports_CSV")
    row = 1
    for filename, header in EXPORT_HEADER_MAP:
        resolved_header = header or _resolve_dynamic_headers(filename)
        if not resolved_header:
            continue
        row = _write_schema_block(
            sheet,
            row,
            f"{filename} (export)",
            resolved_header,
            _build_sample_row(resolved_header),
            "Generado en app.FraudCaseApp._perform_save_exports",
        )

    sheet.cell(row=row, column=1, value="Archivos masivos (entradas CSV)").font = Font(bold=True)
    row += 1
    for label, headers, path in _load_massive_headers(base_dir):
        if not headers:
            continue
        row = _write_schema_block(
            sheet,
            row,
            label,
            headers,
            _build_sample_row(headers),
            f"Fuente: {path.name} leído con models.iter_massive_csv_rows",
        )


def _add_logs_sheet(workbook) -> None:
    sheet = workbook.create_sheet(title="Logs_normalizados")
    sample_log = normalize_log_row(
        {
            "timestamp": "2024-05-01 10:00:00",
            "tipo": "validacion",
            "subtipo": "wireframe",
            "widget_id": "entry_id_caso",
            "coords": (10, 20),
            "mensaje": "Valores de ejemplo normalizados",
            "old_value": "2023-0001",
            "new_value": "2024-0001",
            "action_result": "ok",
        }
    )
    _write_schema_block(
        sheet,
        1,
        "Logs normalizados",
        LOG_FIELDNAMES,
        _build_row_from_mapping(LOG_FIELDNAMES, sample_log),
        "Columnas reales de validators.LOG_FIELDNAMES y validators.normalize_log_row",
    )


def _add_validation_sheet(workbook) -> None:
    sheet = workbook.create_sheet(title="VALIDACIONES")
    headers = ("Hoja", "Campo", "Validador", "Regla")
    for idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=idx, value=header).font = Font(bold=True)
    rows = [
        (
            "Análisis y narrativas",
            "Comentario breve",
            "max_chars",
            str(COMENTARIO_BREVE_MAX_CHARS),
        ),
        (
            "Análisis y narrativas",
            "Comentario breve",
            "allow_newlines",
            "False (sin saltos de línea)",
        ),
        (
            "Análisis y narrativas",
            "Comentario breve",
            "auto_redaccion",
            COMMENTARY_VALIDATIONS["comentario breve"]["auto_redaccion"],
        ),
        (
            "Análisis y narrativas",
            "Comentario amplio",
            "max_chars",
            str(COMENTARIO_AMPLIO_MAX_CHARS),
        ),
        (
            "Análisis y narrativas",
            "Comentario amplio",
            "allow_newlines",
            "False (sin saltos de línea)",
        ),
        (
            "Análisis y narrativas",
            "Comentario amplio",
            "auto_redaccion",
            COMMENTARY_VALIDATIONS["comentario amplio"]["auto_redaccion"],
        ),
        (
            "Caso y participantes",
            "ID Proceso",
            "validate_process_id",
            "Formato BPID-XXXXXX o BPID-RNF-XXXXXX.",
        ),
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)


def _add_supporting_sheets(workbook, base_dir: Path | None = None) -> None:
    base = base_dir or REPO_ROOT
    _add_docx_md_sheet(workbook)
    _add_csv_exports_sheet(workbook, base)
    _add_logs_sheet(workbook)
    _add_validation_sheet(workbook)


def _build_app() -> FraudCaseApp:
    os.environ.setdefault("PYTEST_CURRENT_TEST", "wireframe_export")
    os.environ.setdefault("APP_START_CHOICE", "new")
    try:
        root = tk.Tk()
        root.withdraw()
        root.geometry("1400x900")
    except TclError as exc:  # pragma: no cover - entornos sin display
        raise SystemExit("Tkinter no disponible en el entorno actual") from exc
    return FraudCaseApp(root)


def _build_tooltip_map(app: FraudCaseApp) -> dict[object, str]:
    tooltip_map: dict[object, str] = {}
    for tip in getattr(app, "_hover_tooltips", []) or []:
        widget = getattr(tip, "widget", None)
        text = getattr(tip, "text", None)
        if widget is None or not text:
            continue
        tooltip_map[widget] = str(text)
    return tooltip_map


def _collect_tab_records(app: FraudCaseApp) -> dict[str, list[WidgetRecord]]:
    records: dict[str, list[WidgetRecord]] = {}
    tooltip_map = _build_tooltip_map(app)
    for tab_id in app.notebook.tabs():
        title = app.notebook.tab(tab_id, "text")
        tab_widget = app.notebook.nametowidget(tab_id)
        try:
            app.notebook.select(tab_id)
        except Exception:
            pass
        app.root.update_idletasks()
        origin_x, origin_y = tab_widget.winfo_rootx(), tab_widget.winfo_rooty()
        tab_records = list(
            _iter_widget_records(tab_widget, origin_x, origin_y, (title,), tooltip_map)
        )
        records[title] = tab_records
    return records


def export_wireframes(output_path: Path) -> Path:
    app = _build_app()
    try:
        app.root.update_idletasks()
        tab_records = _collect_tab_records(app)
    finally:
        try:
            app.root.destroy()
        except Exception:
            pass

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for tab_name, records in tab_records.items():
        sheet = wb.create_sheet(title=tab_name[:31])
        next_row = _add_legend(sheet)
        header_row = next_row + 1
        sheet.cell(row=header_row, column=1, value=f"Tab: {tab_name}").font = Font(bold=True)
        _place_records(sheet, records, row_offset=header_row + 1)

    _add_supporting_sheets(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("wireframes/Formulario_UI_wireframe.xlsx"),
        help="Ruta del archivo Excel a generar.",
    )
    args = parser.parse_args()
    output = export_wireframes(args.output)
    print(f"Wireframe exportado en: {output}")


if __name__ == "__main__":  # pragma: no cover - punto de entrada manual
    main()
