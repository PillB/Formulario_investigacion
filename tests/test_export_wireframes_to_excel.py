from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from report_builder import build_event_rows, build_llave_tecnica_rows
from tools.export_wireframes_to_excel import (
    ANALYSIS_HEADERS,
    EXPORT_HEADER_MAP,
    MASSIVE_FILES,
    EMPTY_CASE_DATA,
    _add_supporting_sheets,
    _load_massive_headers,
)
from validators import LOG_FIELDNAMES


REPO_ROOT = Path(__file__).resolve().parents[1]


def _nonempty_rows(sheet):
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        while values and values[-1] is None:
            values.pop()
        if any(value is not None for value in values):
            yield values


def test_supporting_sheets_include_report_and_csv_blocks():
    workbook = Workbook()
    workbook.remove(workbook.active)

    _add_supporting_sheets(workbook, base_dir=REPO_ROOT)

    assert {
        "Reportes_DOCX_MD",
        "Exports_CSV",
        "Logs_normalizados",
    }.issubset(set(workbook.sheetnames))

    report_rows = list(_nonempty_rows(workbook["Reportes_DOCX_MD"]))
    assert build_llave_tecnica_rows(EMPTY_CASE_DATA)[1] in report_rows
    assert build_event_rows(EMPTY_CASE_DATA)[1] in report_rows
    assert list(ANALYSIS_HEADERS) in report_rows

    export_rows = list(_nonempty_rows(workbook["Exports_CSV"]))
    casos_header = list(EXPORT_HEADER_MAP[0][1])  # casos.csv
    assert casos_header in export_rows
    assert LOG_FIELDNAMES in export_rows  # export schema includes logs

    mass_headers = [headers for _, headers, _ in _load_massive_headers(REPO_ROOT) if headers]
    for header in mass_headers:
        assert header in export_rows

    logs_rows = list(_nonempty_rows(workbook["Logs_normalizados"]))
    assert LOG_FIELDNAMES in logs_rows
    log_header_index = logs_rows.index(LOG_FIELDNAMES)
    sample_row = logs_rows[log_header_index + 1]
    assert any("validacion" in str(value).lower() for value in sample_row)


def test_placeholder_rows_capture_formats():
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_supporting_sheets(workbook, base_dir=REPO_ROOT)

    report_rows = list(_nonempty_rows(workbook["Reportes_DOCX_MD"]))
    event_header = build_event_rows(EMPTY_CASE_DATA)[1]
    event_index = report_rows.index(event_header)
    event_sample = report_rows[event_index + 1]
    assert any("YYYY-MM-DD" in str(value) for value in event_sample)
    assert any("2 decimales" in str(value) for value in event_sample)

    export_rows = list(_nonempty_rows(workbook["Exports_CSV"]))
    titles = {row[0] for row in export_rows if len(row) == 1}
    assert any(label in titles for _, label in MASSIVE_FILES)
